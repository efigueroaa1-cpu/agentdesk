"""
core/services/upload_service.py — Ingesta de archivos para análisis (ADR-0003).

Persistencia de uploads con preview de datos estructurados (CSV/Excel),
extraída de core/api.py. Sin FastAPI: recibe bytes + nombre y retorna dicts;
los 404 viajan como LookupError que el borde traduce.
"""
from __future__ import annotations

import json as _json
import logging
import uuid

logger = logging.getLogger(__name__)


def _uploads_dir():
    from core.path_manager import data_path
    return data_path("uploads")


def guardar_upload(nombre_original: str, contenido: bytes) -> dict:
    """Guarda un archivo subido, genera preview CSV/Excel y persiste metadatos."""
    uploads_dir = _uploads_dir()
    uploads_dir.mkdir(parents=True, exist_ok=True)

    archivo_id      = str(uuid.uuid4())[:8]
    nombre_original = nombre_original or "archivo"
    ext             = nombre_original.rsplit(".", 1)[-1].lower() if "." in nombre_original else "bin"
    nombre_interno  = f"{archivo_id}.{ext}"
    ruta            = uploads_dir / nombre_interno
    ruta.write_bytes(contenido)

    # ── Preview de datos estructurados ────────────────────────────────────
    preview: dict = {}
    try:
        if ext == "csv":
            import io, csv
            texto = contenido.decode("utf-8", errors="replace")
            # Detectar separador automáticamente
            dialecto = csv.Sniffer().sniff(texto[:4096], delimiters=",;\t|")
            reader   = csv.DictReader(io.StringIO(texto), dialect=dialecto)
            columnas = reader.fieldnames or []
            filas    = [row for _, row in zip(range(5), reader)]
            preview  = {
                "columnas":     list(columnas),
                "n_columnas":   len(columnas),
                "muestra":      filas,
                "separador":    dialecto.delimiter,
                "total_lineas": texto.count("\n"),
            }
        elif ext in ("xlsx", "xls"):
            import openpyxl
            wb    = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            sheet = wb.active
            filas_raw = list(sheet.iter_rows(min_row=1, max_row=6, values_only=True))
            if filas_raw:
                encabezado = [str(c) if c is not None else "" for c in filas_raw[0]]
                muestra    = [
                    {encabezado[i]: str(v) if v is not None else ""
                     for i, v in enumerate(fila)}
                    for fila in filas_raw[1:]
                ]
                preview = {
                    "columnas":    encabezado,
                    "n_columnas":  len(encabezado),
                    "muestra":     muestra,
                    "hojas":       wb.sheetnames,
                    "total_filas": sheet.max_row,
                }
            wb.close()
    except Exception as exc:
        logger.debug("upload preview error (%s): %s", ext, exc)

    meta = {
        "archivo_id": archivo_id, "nombre_original": nombre_original,
        "nombre_interno": nombre_interno, "tipo": ext,
        "tamano_bytes": len(contenido), "preview": preview,
    }
    (uploads_dir / f"{archivo_id}.meta.json").write_text(
        _json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info("Archivo subido: %s (%d bytes) — %d columnas detectadas",
                nombre_original, len(contenido), len(preview.get("columnas", [])))
    return {
        "archivo_id": archivo_id, "nombre": nombre_original,
        "tipo": ext, "tamano_kb": round(len(contenido) / 1024, 1),
        "preview": preview,
    }


def listar_uploads() -> dict:
    """Lista todos los archivos subidos disponibles para analizar."""
    uploads_dir = _uploads_dir()
    if not uploads_dir.exists():
        return {"archivos": []}
    archivos = []
    for f in sorted(uploads_dir.glob("*.meta.json"), key=lambda x: x.stat().st_mtime, reverse=True):
        try:
            archivos.append(_json.loads(f.read_text(encoding="utf-8")))
        except Exception:
            pass
    return {"archivos": archivos[:30]}


def texto_upload(archivo_id: str) -> dict:
    """Contenido de texto de un archivo subido. LookupError si no existe (404)."""
    uploads_dir = _uploads_dir()
    meta_path   = uploads_dir / f"{archivo_id}.meta.json"
    if not meta_path.exists():
        raise LookupError("Archivo no encontrado.")
    meta    = _json.loads(meta_path.read_text(encoding="utf-8"))
    archivo = uploads_dir / meta["nombre_interno"]
    if not archivo.exists():
        raise LookupError("Contenido no encontrado.")
    texto = archivo.read_bytes().decode("utf-8", errors="replace")[:20_000]
    return {"archivo_id": archivo_id, "nombre": meta["nombre_original"],
            "tipo": meta["tipo"], "texto": texto}
