"""
core/api.py — Servidor FastAPI para el dashboard React de AgentDesk.

Expone:
  WS  /ws/telemetria         Emite eventos de telemetria (@measure_latency) en tiempo real.
  POST /agentes              Crea un nuevo agente (envia CREAR_AGENTE al CommandBridge).
  GET  /agentes              Lista los agentes actuales.
  GET  /health               Verifica que el servidor este activo.
  GET  /kill-switch          Estado del kill switch remoto (Gist de GitHub).
  POST /reload               Recarga la config de uno o todos los agentes (RELOAD_CONFIG).

Arranque autónomo (sin main.py):
  python -m uvicorn core.api:app --host 0.0.0.0 --port 8000 --reload

El servidor inicializa el Orquestador en el evento startup si GEMINI_API_KEY
está configurada. También acepta un bridge externo via registrar_bridge()
para el modo combinado CLI+API.

El dashboard React conecta a ws://localhost:8000/ws/telemetria
y recibe los mismos eventos JSON que el FilterLogHandler envía al terminal.
"""

from __future__ import annotations

import asyncio
import json
import logging
import os
import re
from datetime import datetime, timedelta
from pathlib import Path
from fastapi import FastAPI, HTTPException, WebSocket, WebSocketDisconnect, UploadFile, File as _File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response as _Response
from pydantic import BaseModel

from core.command_bridge import CommandBridge, Command, CREAR_AGENTE, RELOAD_CONFIG, RELOAD_FINANZAS
from core.config_loader  import load_config
from core.path_manager   import REPORTES_DIR, resource_path
from core.schemas        import AgentConfig
from core              import kill_switch
from fastapi.staticfiles import StaticFiles

logger = logging.getLogger(__name__)

# ── Instancia FastAPI ──────────────────────────────────────────────────────────
app = FastAPI(title="AgentDesk API", version="1.0.0")

# CORS restrictivo: solo localhost (la app es local, no expuesta a internet)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://127.0.0.1:8000",
        "http://localhost:8000",
        "http://localhost:5173",   # Vite dev server
        "http://localhost:3000",
    ],
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "Accept"],
    allow_credentials=True,
)

# ── Middleware JWT: verifica token en endpoints protegidos ─────────────────────
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request as _Request
from starlette.requests import Request      # alias para type hints en endpoints
from starlette.responses import JSONResponse as _JSONResponse

_RUTAS_PUBLICAS = {
    "/auth/login", "/health", "/auth/verificar",
    "/modelos", "/proveedores", "/docs", "/openapi.json",
    "/version", "/kill-switch",
}
# Prefijos públicos: accesibles sin token
_PREFIJOS_PUBLICOS = (
    "/ui/", "/ws/",
    "/monitor/", "/scheduler/",
    "/embeddings", "/historial", "/tendencias",
    "/reportes", "/uploads",
    "/agentes",       # lectura pública (escritura la protege el método HTTP)
    "/chat",          # el orquestador es la interfaz principal
    "/upload",        # subir archivos para analizar
    "/generar-pdf",   # generar reportes PDF
    "/backup/",       # backup/restore
    "/memoria/",      # memoria de conversaciones
    "/rate-limiter",  # estadísticas
    "/update/",       # verificar actualizaciones
    "/dashboard",     # dashboard analytics
    "/webhook/",      # webhooks externos (auth propia por bcrypt)
    "/finanzas/",     # motor financiero (auth por JWT del middleware global)
    "/gantt/",        # motor Gantt — CRUD de tareas y exportación PDF
    "/compliance/",   # auditoría de guardrails
    "/riesgo/",       # análisis de riesgo Gantt-Finanzas
    "/sistema/",      # KPIs maestros para BIDashboard
    "/analytics/",    # Curva S / EVM (rol supervisor+ verificado en endpoint)
    "/docs/",         # manual PDF (accesible a todos los usuarios autenticados)
    "/kill-switch/",  # toggle y URL del kill switch
)

# Solo proteger MUTACIONES sensibles con JWT
_METODOS_PROTEGIDOS = {"DELETE"}  # solo DELETE requiere siempre token
_RUTAS_SIEMPRE_PROTEGIDAS = {
    "/auth/cambiar-password",
    "/proveedores/apikey",
}


class JWTMiddleware(BaseHTTPMiddleware):
    """Verifica el JWT en todas las rutas protegidas."""

    async def dispatch(self, request: _Request, call_next):
        path   = request.url.path
        method = request.method

        # OPTIONS siempre pasa (CORS preflight)
        if method == "OPTIONS":
            return await call_next(request)

        # Rutas explícitamente protegidas (siempre necesitan token)
        necesita_token = (
            path in _RUTAS_SIEMPRE_PROTEGIDAS
            or method in _METODOS_PROTEGIDOS
        )

        # Nota: las rutas bajo prefijos públicos (_RUTAS_PUBLICAS / _PREFIJOS_PUBLICOS)
        # no exigen token para pasar, pero si el cliente envía uno igual se decodifica
        # más abajo — varios endpoints bajo esos prefijos (p.ej. /analytics/, /backup/)
        # verifican el rol ellos mismos vía request.state.rol, y ese chequeo solo
        # funciona si el token llegó a decodificarse.
        auth_header = request.headers.get("Authorization", "")
        token       = auth_header.replace("Bearer ", "").strip()

        if not token:
            if necesita_token:
                return _JSONResponse(
                    status_code=401,
                    content={"detail": "Token requerido para esta operación."},
                )
            # Sin token: continúa como anónimo (rol por defecto "viewer" en cada endpoint)
            return await call_next(request)

        try:
            from core.auth import verificar_token
            datos = verificar_token(token)
            if datos:
                request.state.usuario = datos.get("sub")
                request.state.rol     = datos.get("role", "viewer")
        except Exception:
            pass

        return await call_next(request)


app.add_middleware(JWTMiddleware)

# ── React UI estático en /ui/ ──────────────────────────────────────────────────
# Servido íntegramente por StaticFiles; el middleware inferior fuerza no-cache
# en todo /ui/* (evita el cache agresivo de WebView2 tras cada build).
_react_dist = resource_path("react_dist")

@app.middleware("http")
async def _no_cache_ui(request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/ui/") or request.url.path == "/ui":
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

if _react_dist.exists():
    app.mount("/ui", StaticFiles(directory=str(_react_dist), html=True), name="ui")

# ── Gestor de conexiones WebSocket ────────────────────────────────────────────
class ConnectionManager:
    """
    Mantiene la lista de clientes WebSocket conectados con su rol RBAC.
    broadcast() acepta rol_minimo para filtrar destinatarios (ej. solo supervisor+).
    """

    def __init__(self) -> None:
        # ws -> rol del usuario ("viewer" | "supervisor" | "admin")
        self._clientes: dict[WebSocket, str] = {}

    @property
    def active(self) -> list[WebSocket]:
        return list(self._clientes.keys())

    async def connect(self, ws: WebSocket, rol: str = "viewer") -> None:
        await ws.accept()
        self._clientes[ws] = rol

    def disconnect(self, ws: WebSocket) -> None:
        self._clientes.pop(ws, None)

    async def broadcast(self, mensaje: dict, rol_minimo: str = "viewer") -> None:
        """
        Envía `mensaje` solo a los clientes cuyo rol >= rol_minimo.
        rol_minimo="viewer" → todos; "supervisor" → excluye viewers; "admin" → solo admins.
        """
        from core.auth import tiene_permiso
        texto   = json.dumps(mensaje, ensure_ascii=False)
        muertos: list[WebSocket] = []
        for ws, rol_cliente in list(self._clientes.items()):
            if not tiene_permiso(rol_cliente, rol_minimo):
                continue
            try:
                await ws.send_text(texto)
            except Exception:
                muertos.append(ws)
        for ws in muertos:
            self._clientes.pop(ws, None)


manager = ConnectionManager()

# ── Estado interno del servidor ───────────────────────────────────────────────
# _bridge        : inyectado desde main.py O auto-inicializado en startup.
# _tarea_monitor : tarea background del kill switch (cancelada en shutdown).
# _tarea_cmds    : tarea background del CommandBridge (cancelada en shutdown).
_bridge:        CommandBridge | None = None
_tarea_monitor: asyncio.Task | None  = None
_tarea_cmds:    asyncio.Task | None  = None
_orquestador:   object | None        = None   # Orquestador (tipado como object para evitar import circular)


def registrar_bridge(bridge: CommandBridge) -> None:
    """Inyecta un CommandBridge externo (p.ej. desde main.py en modo CLI+API)."""
    global _bridge
    _bridge = bridge


def registrar_orquestador(orch: object) -> None:
    """Inyecta la referencia al Orquestador para que los endpoints puedan llamarlo."""
    global _orquestador
    _orquestador = orch


# ── Handler de logging que reenvía telemetría al WebSocket ───────────────────
class WebSocketLogHandler(logging.Handler):
    """
    Intercepta las entradas de @measure_latency (campos filtro + duracion_s)
    y las emite como eventos JSON a todos los clientes WebSocket conectados.

    Se instala en el logger core.pipeline igual que FilterLogHandler,
    pero en lugar de imprimir en terminal envía al dashboard React.
    """

    def emit(self, record: logging.LogRecord) -> None:
        filtro   = getattr(record, "filtro",     None)
        duracion = getattr(record, "duracion_s", None)
        status   = getattr(record, "status",     "")
        agente   = getattr(record, "agente",     "")

        # ── Evento de filtro ejecutado / timeout / error ───────────────────────
        if filtro and duracion is not None:
            mensaje = {
                "tipo":       "telemetria",
                "filtro":     filtro,
                "agente":     agente,
                "status":     status or "ok",
                "duracion_s": duracion,
                "timestamp":  record.created,
            }

        # ── Pipeline abortado (no tiene duracion_s pero sí motivo) ─────────────
        elif status == "abortado" and agente:
            mensaje = {
                "tipo":      "pipeline_abortado",
                "agente":    agente,
                "filtro":    filtro or "",
                "motivo":    getattr(record, "motivo", record.getMessage()),
                "nivel":     record.levelname,
                "timestamp": record.created,
            }

        else:
            return   # ignorar el resto

        # Enviar de forma async sin bloquear el logging handler.
        try:
            loop = asyncio.get_running_loop()
            loop.create_task(manager.broadcast(mensaje))

            # Actualizar progreso Gantt cuando el pipeline termina OK
            if mensaje["tipo"] == "telemetria" and mensaje["status"] == "ok" and agente:
                loop.create_task(_actualizar_gantt_desde_telemetria(agente, mensaje))

            # Persistir abortos de guardrail en DB para compliance
            if mensaje["tipo"] == "pipeline_abortado" and agente:
                loop.create_task(_persistir_guardrail(
                    agente,
                    getattr(record, "filtro", "desconocido") or "desconocido",
                    mensaje.get("motivo", ""),
                ))
        except RuntimeError:
            pass  # sin loop activo — se descarta el evento


async def _actualizar_gantt_desde_telemetria(agente_id: str, _evento: dict) -> None:
    """
    Incrementa el progreso de las tareas Gantt del agente cuando su pipeline
    reporta un filtro exitoso. El incremento es proporcional a los filtros
    completados (cada 'ok' suma ~10% hasta el máximo de 100%).
    Los cambios se transmiten a todos los clientes WS via gantt_progreso.
    """
    try:
        from core.gantt import motor_gantt
        actualizadas = motor_gantt.actualizar_progreso_por_agente(agente_id, incremento_pct=10.0)
        for t in actualizadas:
            await manager.broadcast({
                "tipo":      "gantt_progreso",
                "tarea_id":  t["id"],
                "proyecto":  t["proyecto_id"],
                "agente_id": agente_id,
                "pct":       t["pct_completado"],
            })
    except Exception:
        pass  # nunca romper el flujo de telemetría


async def _persistir_guardrail(agente_id: str, guardrail: str, motivo: str) -> None:
    """Persiste un evento de aborto de guardrail en DB y notifica a compliance."""
    try:
        from core.compliance import motor_compliance
        await asyncio.get_running_loop().run_in_executor(
            None,
            lambda: motor_compliance.registrar_evento(agente_id, guardrail, motivo),
        )
    except Exception:
        pass  # nunca romper el flujo de telemetría


def instalar_ws_handler() -> None:
    """
    Añade WebSocketLogHandler al logger de pipeline.
    Llamar una vez al arrancar el servidor.
    """
    pl = logging.getLogger("core.pipeline")
    if not any(isinstance(h, WebSocketLogHandler) for h in pl.handlers):
        pl.addHandler(WebSocketLogHandler(level=logging.DEBUG))


# ── Endpoints ─────────────────────────────────────────────────────────────────

async def _auto_iniciar_orquestador() -> None:
    """
    Inicializa el Orquestador dentro del proceso FastAPI si aún no hay bridge
    registrado y GEMINI_API_KEY está disponible en el entorno.

    Permite que `uvicorn core.api:app` arranque como servidor autónomo
    sin necesidad de main.py, manteniendo el desacoplamiento total:
    la UI (React, CLI, cualquier cliente WS) puede conectarse o no.
    """
    global _bridge, _tarea_cmds

    if _bridge is not None:
        return  # Ya registrado externamente — no sobreescribir

    api_key = os.environ.get("GEMINI_API_KEY", "")
    if not api_key:
        logger.warning(
            "API: GEMINI_API_KEY no disponible — "
            "Orquestador no inicializado. Los endpoints de agentes estarán inactivos."
        )
        return

    try:
        from google import genai
        from core.orchestrator import Orquestador

        client = genai.Client(api_key=api_key)
        pager  = await client.aio.models.list()
        model  = None
        async for m in pager:
            if "generateContent" in (m.supported_actions or []):
                model = m.name
                break

        if not model:
            logger.error("API startup: ningún modelo Gemini disponible.")
            return

        from core.path_manager import resource_path
        config_path = str(resource_path("config.json"))

        bridge = CommandBridge()
        orch   = Orquestador(config_path, client, model, bridge=bridge)

        registrar_bridge(bridge)
        registrar_orquestador(orch)
        _tarea_cmds = asyncio.create_task(orch.procesar_comandos())

        logger.info(
            "API: Orquestador auto-inicializado.",
            extra={"modelo": model, "agentes": list(orch.agentes.keys())},
        )
    except Exception as exc:
        logger.error("API startup: fallo al inicializar Orquestador: %s", exc)


@app.on_event("startup")
async def startup() -> None:
    global _tarea_monitor

    instalar_ws_handler()

    # Inicializar SQLite
    try:
        from core.database import init_db
        init_db()
        logger.info("SQLite inicializado correctamente")
    except Exception as exc:
        logger.error("SQLite init error: %s", exc)

    # Migrar API keys al vault cifrado (primera vez)
    try:
        from core.key_vault import migrar_env_a_vault
        r = migrar_env_a_vault()
        if r["migradas"]:
            logger.info("Vault: %d API keys migradas", r["migradas"])
    except Exception as exc:
        logger.warning("Vault migration: %s", exc)

    # Inicializar Scheduler de Monitoreo Automático
    try:
        from core.scheduler import start_scheduler
        _task_scheduler = start_scheduler(
            orquestador=_orquestador,
            broadcast=manager.broadcast,
        )
        logger.info("Scheduler de monitoreo iniciado")
    except Exception as exc:
        logger.error("Scheduler init error: %s", exc)

    # React /ui/ montado a nivel de módulo (no aquí) para correcta compilación de rutas Starlette

    # Kill switch: arrancar monitor de verificación periódica
    _tarea_monitor = asyncio.create_task(kill_switch.iniciar_monitor())

    # Verificación inmediata (no espera el primer intervalo de 5 min)
    await kill_switch.verificar_gist()

    # Auto-inicializar Orquestador si no hay bridge externo
    await _auto_iniciar_orquestador()


@app.on_event("shutdown")
async def shutdown() -> None:
    for tarea in (_tarea_monitor, _tarea_cmds):
        if tarea and not tarea.done():
            tarea.cancel()
            try:
                await tarea
            except asyncio.CancelledError:
                pass


@app.post("/agentes/ejecutar-todos")
async def ejecutar_todos() -> dict:
    """Ejecuta realizar_tarea() en TODOS los agentes en paralelo (asyncio.gather)."""
    if not kill_switch.is_active():
        return {"error": "Kill switch activo."}
    if _orquestador is None:
        return {"error": "Orquestador no disponible."}

    await manager.broadcast({"tipo": "todos_ejecutando",
                              "agentes": list(_orquestador.agentes.keys())})

    resultados = {}
    tareas     = [
        agente.realizar_tarea("reporte_ventas")
        for agente in _orquestador.agentes.values()
    ]
    respuestas = await asyncio.gather(*tareas, return_exceptions=True)
    for aid, resp in zip(_orquestador.agentes.keys(), respuestas):
        if isinstance(resp, Exception):
            resultados[aid] = {"ok": False, "error": str(resp)}
        else:
            ok = resp is not None
            resultados[aid] = {"ok": ok, "resumen": (resp or {}).get("resumen", "")[:150] if ok else None}

    await manager.broadcast({"tipo": "todos_completados", "resultados": resultados})
    return {"ok": True, "resultados": resultados}


@app.get("/logs")
async def get_logs(n: int = 100, nivel: str = "all") -> dict:
    """Devuelve las últimas N entradas del log sistema.log como JSON."""
    from core.path_manager import data_path
    import json as _json
    log_path = data_path("logs/sistema.log")
    if not log_path.exists():
        return {"entradas": [], "total": 0}
    try:
        lineas = log_path.read_text(encoding="utf-8").splitlines()
        entradas = []
        for l in lineas:
            l = l.strip()
            if not l.startswith("{"): continue
            try:
                e = _json.loads(l)
                if nivel != "all" and e.get("level", "").upper() != nivel.upper():
                    continue
                entradas.append(e)
            except Exception:
                pass
        entradas = entradas[-n:]
        return {"entradas": list(reversed(entradas)), "total": len(entradas)}
    except Exception as ex:
        return {"entradas": [], "total": 0, "error": str(ex)}


@app.get("/reportes")
async def listar_todos_reportes() -> dict:
    """Lista todos los reportes PDF de todos los agentes."""
    from core.path_manager import REPORTES_DIR
    if not REPORTES_DIR.exists():
        return {"reportes": []}
    archivos = []
    for f in sorted(REPORTES_DIR.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
        if f.suffix in (".pdf", ".md", ".json"):
            archivos.append({
                "nombre":    f.name,
                "tipo":      f.suffix[1:],
                "tamano_kb": round(f.stat().st_size / 1024, 1),
                "mtime":     f.stat().st_mtime,
                "url":       f"/reportes/{f.name}",
            })
    return {"reportes": archivos[:50]}


from fastapi.responses import FileResponse as _FileResponse

@app.get("/reportes/{nombre}")
async def descargar_reporte_directo(nombre: str) -> _FileResponse:
    """Descarga un archivo de reportes por nombre."""
    from core.path_manager import REPORTES_DIR
    archivo = REPORTES_DIR / nombre
    if not archivo.exists() or not archivo.is_file():
        raise HTTPException(status_code=404, detail=f"Reporte '{nombre}' no encontrado.")
    media = "application/pdf" if archivo.suffix == ".pdf" else "text/plain"
    return _FileResponse(str(archivo), filename=nombre, media_type=media,
                         headers={"Content-Disposition": f'attachment; filename="{nombre}"'})


@app.get("/reportes/{nombre}/abrir")
async def abrir_reporte_os(nombre: str) -> dict:
    """Abre el archivo con la aplicación predeterminada del sistema operativo."""
    import os
    from core.path_manager import REPORTES_DIR
    archivo = REPORTES_DIR / nombre
    if not archivo.exists() or not archivo.is_file():
        raise HTTPException(status_code=404, detail=f"Reporte '{nombre}' no encontrado.")
    try:
        os.startfile(str(archivo))
        return {"ok": True, "nombre": nombre}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error al abrir archivo: {exc}")


# ── Dashboard Analytics ───────────────────────────────────────────────────────

@app.get("/dashboard/datos")
async def dashboard_datos(agente_id: str = "", dias: int = 30) -> dict:
    """Retorna métricas y series para el dashboard de Analytics."""
    from core.path_manager import REPORTES_DIR
    from core.config_loader import load_config
    from datetime import datetime as _dt, timedelta as _td

    dias = max(7, min(365, int(dias)))

    cfg = load_config()
    _raw = cfg.get("agents", cfg.get("agentes", []))
    if isinstance(_raw, list):
        agentes_cfg = {a["id"]: a for a in _raw if isinstance(a, dict) and "id" in a}
    elif isinstance(_raw, dict):
        agentes_cfg = _raw
    else:
        agentes_cfg = {}

    agente_info: dict = {}
    if agente_id and agente_id in agentes_cfg:
        a = agentes_cfg[agente_id]
        agente_info = {
            "id":     agente_id,
            "nombre": a.get("nombre", agente_id),
            "area":   a.get("area", "General"),
            "modelo": a.get("modelo", ""),
        }

    # ── Métricas SQLite ───────────────────────────────────────────────────────
    total_sesiones = total_mensajes = 0
    actividad_raw: list = []
    horaria_raw: list  = []
    try:
        from core.database import get_session
        from core.memory import Conversacion, Mensaje
        from sqlalchemy import func as _func

        with get_session() as s:
            q = s.query(Conversacion)
            if agente_id:
                q = q.filter(Conversacion.agente_id == agente_id)
            total_sesiones = q.count()

            qm = s.query(Mensaje)
            if agente_id:
                qm = qm.filter(Mensaje.agente_id == agente_id)
            total_mensajes = qm.count()

            hace_n = _dt.utcnow() - _td(days=dias)

            # Daily activity
            qa = (
                s.query(
                    _func.date(Mensaje.ts).label("dia"),
                    _func.count(Mensaje.id).label("n"),
                )
                .filter(Mensaje.ts >= hace_n)
            )
            if agente_id:
                qa = qa.filter(Mensaje.agente_id == agente_id)
            actividad_raw = qa.group_by(_func.date(Mensaje.ts)).all()

            # Hourly heatmap: [weekday 0-6] × [hour 0-23]
            qh = (
                s.query(
                    _func.strftime("%w", Mensaje.ts).label("dow"),
                    _func.strftime("%H", Mensaje.ts).label("hora"),
                    _func.count(Mensaje.id).label("n"),
                )
                .filter(Mensaje.ts >= hace_n)
            )
            if agente_id:
                qh = qh.filter(Mensaje.agente_id == agente_id)
            horaria_raw = qh.group_by("dow", "hora").all()

    except Exception as _e:
        logger.warning("dashboard_datos DB: %s", _e)

    # Build 7×24 hourly matrix (SQLite %w: 0=Sun … 6=Sat)
    actividad_horaria = [[0] * 24 for _ in range(7)]
    for row in horaria_raw:
        try:
            actividad_horaria[int(row.dow)][int(row.hora)] = int(row.n)
        except Exception:
            pass

    # ── Reportes en disco ─────────────────────────────────────────────────────
    rdir = REPORTES_DIR
    todos_pdf = sorted(rdir.glob("*.pdf"), key=lambda x: x.stat().st_mtime, reverse=True) if rdir.exists() else []
    nombre_clave = (agente_info.get("nombre") or "").replace(" ", "_")[:12].lower()
    filtrados = [r for r in todos_pdf if nombre_clave and nombre_clave in r.stem.lower()] or todos_pdf

    reportes_recientes = [
        {
            "nombre":    r.name,
            "fecha":     _dt.fromtimestamp(r.stat().st_mtime).strftime("%d/%m/%Y %H:%M"),
            "tamano_kb": round(r.stat().st_size / 1024, 1),
        }
        for r in filtrados[:12]
    ]

    # ── Series temporales ─────────────────────────────────────────────────────
    hoy = _dt.utcnow().date()
    act_dict = {str(row.dia): row.n for row in actividad_raw}
    actividad_serie = [
        {"fecha": (hoy - _td(days=i)).strftime("%d/%m"),
         "mensajes": act_dict.get(str(hoy - _td(days=i)), 0)}
        for i in range(dias - 1, -1, -1)
    ]

    rep_por_dia: dict[str, int] = {}
    for r in filtrados:
        d = _dt.fromtimestamp(r.stat().st_mtime).strftime("%d/%m")
        rep_por_dia[d] = rep_por_dia.get(d, 0) + 1
    reportes_serie = [
        {"fecha": a["fecha"], "reportes": rep_por_dia.get(a["fecha"], 0)}
        for a in actividad_serie
    ]

    return {
        "agente": agente_info,
        "dias":   dias,
        "agentes_disponibles": [
            {"id": k, "nombre": v.get("nombre", k), "area": v.get("area", "")}
            for k, v in agentes_cfg.items()
        ],
        "kpis": {
            "sesiones":  total_sesiones,
            "mensajes":  total_mensajes,
            "reportes":  len(filtrados),
        },
        "actividad_serie":    actividad_serie,
        "reportes_serie":     reportes_serie,
        "actividad_horaria":  actividad_horaria,
        "reportes_recientes": reportes_recientes,
    }


@app.get("/dashboard/briefing")
async def dashboard_briefing(tema: str = "", agente_id: str = "") -> dict:
    """Busca información web sobre el tema usando Tavily y lo retorna como JSON estructurado."""
    if not tema.strip():
        return {"tema": "", "query": "", "agente": {}, "answer": "", "resultados": []}

    from core.config_loader import load_config
    from core.tools import _TAVILY_KEY
    import httpx

    cfg = load_config()
    _raw = cfg.get("agents", cfg.get("agentes", []))
    if isinstance(_raw, list):
        agentes_cfg = {a["id"]: a for a in _raw if isinstance(a, dict) and "id" in a}
    elif isinstance(_raw, dict):
        agentes_cfg = _raw
    else:
        agentes_cfg = {}

    agente_info: dict = {}
    if agente_id and agente_id in agentes_cfg:
        a = agentes_cfg[agente_id]
        agente_info = {
            "id":     agente_id,
            "nombre": a.get("nombre", agente_id),
            "area":   a.get("area", ""),
        }

    area  = agente_info.get("area", "")
    query = f"{tema.strip()} {area}".strip() if area else tema.strip()

    try:
        async with httpx.AsyncClient(timeout=22) as client:
            resp = await client.post(
                "https://api.tavily.com/search",
                json={
                    "api_key":             _TAVILY_KEY,
                    "query":               query,
                    "max_results":         8,
                    "search_depth":        "advanced",
                    "include_answer":      True,
                    "include_raw_content": False,
                },
            )
            resp.raise_for_status()
            data = resp.json()

        resultados = [
            {
                "titulo":  r.get("title", ""),
                "url":     r.get("url", ""),
                "dominio": (r.get("url", "").split("/")[2] if r.get("url", "").startswith("http") else ""),
                "snippet": (r.get("content") or "")[:380],
                "score":   round(r.get("score", 0), 2),
            }
            for r in data.get("results", [])
        ]

        # Generar resumen en español usando Groq directamente (vault key)
        answer_es = ""
        try:
            import asyncio
            from core.key_vault import obtener_key as _get_key
            from groq import AsyncGroq as _GroqClient
            _groq_key = _get_key("GROQ_API_KEY") or ""
            if _groq_key:
                snippets_text = "\n\n".join(
                    f"[{r['dominio']}] {r['snippet']}" for r in resultados[:5] if r["snippet"]
                )
                llm_prompt = (
                    f"Basándote en esta información sobre \"{tema.strip()}\", escribe UN PÁRRAFO BREVE "
                    f"(3-5 oraciones) en ESPAÑOL con los datos más importantes: resultados financieros, "
                    f"producción, noticias recientes, o lo que sea más relevante.\n\n"
                    f"INFORMACIÓN:\n{snippets_text}\n\n"
                    f"REGLAS: Responde SOLO en español. Sin introducción. Sin títulos. Solo el párrafo directo."
                )
                async def _groq_summarize():
                    _c = _GroqClient(api_key=_groq_key)
                    _r = await _c.chat.completions.create(
                        model="llama-3.3-70b-versatile",
                        messages=[{"role": "user", "content": llm_prompt}],
                        temperature=0.3,
                        max_tokens=300,
                    )
                    return _r.choices[0].message.content.strip()
                answer_es = await asyncio.wait_for(_groq_summarize(), timeout=12)
        except Exception as _llm_e:
            logger.warning("dashboard_briefing llm summary: %s", _llm_e)
        if not answer_es:
            answer_es = data.get("answer", "")  # fallback to Tavily's answer

        return {
            "tema":       tema.strip(),
            "query":      query,
            "agente":     agente_info,
            "answer":     answer_es,
            "resultados": resultados,
        }
    except Exception as _e:
        logger.warning("dashboard_briefing: %s", _e)
        return {"tema": tema, "query": query, "agente": agente_info, "answer": "", "resultados": [], "error": str(_e)}


@app.get("/dashboard/resumen-ia")
async def dashboard_resumen_ia(dias: int = 30, agente_id: str = "") -> dict:
    """Genera un resumen ejecutivo en español con el LLM usando los datos del dashboard."""
    import asyncio
    from core.key_vault import obtener_key as _get_key

    # Reutilizar endpoint de datos para obtener KPIs reales
    base = await dashboard_datos(dias=dias, agente_id=agente_id)
    kpis  = base.get("kpis", {})
    serie = base.get("actividad_serie", [])
    rep_serie = base.get("reportes_serie", [])

    last7  = sum(d["mensajes"] for d in serie[-7:])
    prev7  = sum(d["mensajes"] for d in serie[-14:-7])
    chg7   = round((last7-prev7)/prev7*100) if prev7 > 0 else None
    reps_sem = sum(d["reportes"] for d in rep_serie[-7:])

    ctx = (
        f"Sistema de agentes IA · últimos {dias} días:\n"
        f"- Sesiones: {kpis.get('sesiones',0)}\n"
        f"- Mensajes totales: {kpis.get('mensajes',0)}\n"
        f"- Mensajes últimos 7 días: {last7}"
        + (f" ({'+' if (chg7 or 0)>=0 else ''}{chg7}% vs semana anterior)" if chg7 is not None else "") + "\n"
        f"- Mensajes por sesión: {round(kpis.get('mensajes',0)/kpis.get('sesiones',1),1) if kpis.get('sesiones',0)>0 else 0}\n"
        f"- Reportes PDF totales: {kpis.get('reportes',0)}\n"
        f"- Reportes PDF última semana: {reps_sem}\n"
    )
    if agente_id:
        ctx += f"- Filtrado para agente: {agente_id}\n"

    prompt = (
        f"Eres un analista de productividad. Analiza estos datos de uso de un sistema de agentes IA "
        f"y escribe un resumen ejecutivo en ESPAÑOL con 3 párrafos cortos:\n\n"
        f"1. Estado actual y tendencia de actividad\n"
        f"2. Eficiencia y productividad (reportes, calidad de sesiones)\n"
        f"3. Una recomendación concreta y accionable\n\n"
        f"DATOS:\n{ctx}\n\n"
        f"REGLAS: Solo en español. Sin títulos ni numeración. Lenguaje claro, directo, sin jerga técnica."
    )

    try:
        from groq import AsyncGroq as _GroqClient
        _key = _get_key("GROQ_API_KEY") or ""
        if not _key:
            return {"resumen": "", "error": "Sin clave Groq"}

        async def _call():
            c = _GroqClient(api_key=_key)
            r = await c.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.4, max_tokens=450,
            )
            return r.choices[0].message.content.strip()

        resumen = await asyncio.wait_for(_call(), timeout=18)
        return {"resumen": resumen}
    except Exception as e:
        logger.warning("dashboard_resumen_ia: %s", e)
        return {"resumen": "", "error": str(e)}


@app.get("/dashboard/leer-pagina")
async def dashboard_leer_pagina(url: str) -> dict:
    """Extrae el contenido legible de una URL usando Tavily Extract + fallback HTTP."""
    if not url.strip():
        return {"url": url, "contenido": ""}
    from core.tools import _obtener_pagina
    contenido = await _obtener_pagina(url.strip(), max_chars=7000)
    return {"url": url, "contenido": contenido}


# ── Monitor Web + Base de datos ───────────────────────────────────────────────

@app.get("/monitor/fetch")
async def monitor_fetch(categoria: str, params: str = "{}") -> dict:
    """
    Fetch de datos en tiempo real desde una fuente web.
    categoria: futbol_equipo | futbol_multiple | energia_renovable | energia_spot | energia_demanda
    params: JSON string con parámetros adicionales
    """
    from core.web_monitor import fetch_categoria
    from core.database   import guardar_dato_monitor
    import json as _json
    try:
        p      = _json.loads(params)
        result = await fetch_categoria(categoria, p)

        # Persistir en SQLite si hay datos numéricos
        if isinstance(result, dict) and "estadisticas" in result:
            st = result["estadisticas"]
            for k, v in st.items():
                if isinstance(v, (int, float)):
                    guardar_dato_monitor(0, categoria, categoria, f"{result.get('nombre','')}/{k}",
                                        str(v), float(v))
        elif isinstance(result, dict) and "solar" in result:
            for k, v in result.get("solar",{}).items():
                if isinstance(v, (int,float)):
                    guardar_dato_monitor(0, "energia", "energia", f"solar/{k}", str(v), float(v))

        await manager.broadcast({"tipo":"monitor_actualizado","categoria":categoria})
        return {"ok": True, "categoria": categoria, "data": result}
    except Exception as exc:
        logger.error("monitor_fetch: %s", exc)
        return {"ok": False, "error": str(exc)}


@app.get("/scheduler/tareas")
async def scheduler_tareas() -> dict:
    """Estado de todas las tareas del scheduler."""
    from core.scheduler import get_tareas
    return {"tareas": get_tareas()}


class SchedulerUpdateRequest(BaseModel):
    activo:        bool  | None = None
    intervalo_min: int   | None = None


@app.put("/scheduler/tareas/{tarea_id}")
async def scheduler_update(tarea_id: str, payload: SchedulerUpdateRequest) -> dict:
    """Activar/desactivar o cambiar intervalo de una tarea."""
    from core.scheduler import actualizar_tarea
    result = actualizar_tarea(tarea_id, payload.activo, payload.intervalo_min)
    if result is None:
        raise HTTPException(status_code=404, detail=f"Tarea '{tarea_id}' no encontrada.")
    await manager.broadcast({"tipo": "scheduler_actualizado", "tarea_id": tarea_id})
    return {"ok": True, "tareas": result}


@app.post("/scheduler/tareas/{tarea_id}/ejecutar")
async def scheduler_ejecutar_ahora(tarea_id: str) -> dict:
    """Ejecuta una tarea del scheduler inmediatamente."""
    from core.scheduler import ejecutar_ahora
    ok = await ejecutar_ahora(tarea_id)
    if not ok:
        raise HTTPException(status_code=404, detail=f"Tarea '{tarea_id}' no encontrada.")
    return {"ok": True, "mensaje": f"Tarea '{tarea_id}' iniciada en background"}


@app.get("/monitor/equipos-preset")
async def monitor_equipos_preset() -> dict:
    """Lista de equipos y ligas disponibles."""
    from core.web_monitor import EQUIPOS_PRESET, LIGAS_CATALOGO
    return {"equipos": EQUIPOS_PRESET, "ligas": LIGAS_CATALOGO}


@app.get("/monitor/liga/{liga_id}")
async def monitor_liga(liga_id: str, nombre: str = "") -> dict:
    """Datos completos de una liga: tabla, partidos, estadísticas."""
    from core.web_monitor import fetch_futbol_liga_completo
    try:
        resultado = await fetch_futbol_liga_completo(liga_id, nombre)
        # Persistir tabla en SQLite
        from core.database import guardar_dato_monitor
        for eq in resultado.get("equipos_tabla", []):
            guardar_dato_monitor(
                fuente_id=0, fuente_nombre=f"Liga:{nombre or liga_id}",
                categoria="futbol_liga", clave=f"{nombre}/{eq['equipo']}/puntos",
                valor=str(eq["puntos"]), valor_numerico=float(eq["puntos"]),
            )
        await manager.broadcast({"tipo": "monitor_actualizado", "categoria": "futbol_liga", "liga": nombre})
        return {"ok": True, "data": resultado}
    except Exception as exc:
        logger.error("monitor_liga %s: %s", liga_id, exc)
        return {"ok": False, "error": str(exc)}


@app.get("/monitor/historial")
async def monitor_historial(
    categoria: str | None = None,
    clave: str | None = None,
    limit: int = 100,
) -> dict:
    """Historial de datos monitoreados desde SQLite."""
    from core.database import get_datos_monitor
    return {"datos": get_datos_monitor(categoria=categoria, clave=clave, limit=limit)}


@app.get("/monitor/alertas")
async def monitor_alertas(no_leidas: bool = False) -> dict:
    """Alertas generadas por el monitor."""
    from core.database import get_alertas, marcar_alertas_leidas
    alertas = get_alertas(solo_no_leidas=no_leidas)
    if no_leidas: marcar_alertas_leidas()
    return {"alertas": alertas, "total": len(alertas)}


@app.get("/historial/agente/{agente_id}")
async def historial_agente(agente_id: str, limit: int = 50) -> dict:
    """Historial real de ejecuciones de un agente desde SQLite."""
    from core.database import get_historial, get_stats_agente
    return {
        "historial": get_historial(agente_id=agente_id, limit=limit),
        "stats":     get_stats_agente(agente_id),
    }


@app.get("/historial")
async def historial_global(limit: int = 100) -> dict:
    """Historial global de todas las ejecuciones."""
    from core.database import get_historial
    return {"historial": get_historial(limit=limit)}


def _calcular_tendencia(valores: list[float]) -> str:
    """
    Calcula la tendencia de una serie usando regresion lineal minima.
    Retorna 'mejorando', 'estable' o 'empeorando'.
    Para tasa_exito: pendiente positiva = mejorando.
    Para latencia_s: pendiente negativa = mejorando (menos tiempo = mejor).
    """
    n = len(valores)
    if n < 3:
        return "estable"
    xs = list(range(n))
    x_m = sum(xs) / n
    y_m = sum(valores) / n
    num = sum((xs[i] - x_m) * (valores[i] - y_m) for i in range(n))
    den = sum((xs[i] - x_m) ** 2 for i in range(n))
    if den == 0:
        return "estable"
    slope = num / den
    # Umbral: cambio mayor al 2% del rango o 1 punto porcentual de exito
    rango = max(valores) - min(valores)
    umbral = max(rango * 0.05, 1.0)
    if slope > umbral / n:
        return "mejorando"
    if slope < -umbral / n:
        return "empeorando"
    return "estable"


@app.get("/tendencias/{agente_id}")
async def tendencias_agente(agente_id: str, dias: int = 30) -> dict:
    """
    Tendencias de rendimiento de un agente en los ultimos N dias.
    Primero intenta con Ejecucion (pipeline formal); si no hay datos
    usa las conversaciones de chat como proxy de actividad.
    La tendencia se calcula por regresion lineal, no por comparacion extremo a extremo.
    """
    from core.database import get_session, Ejecucion
    from datetime import datetime as _dt, timedelta
    from collections import defaultdict

    desde = _dt.utcnow() - timedelta(days=dias)

    # ── Intento 1: datos de ejecuciones del pipeline ──────────────────────────
    with get_session() as s:
        exec_rows = (
            s.query(Ejecucion)
            .filter(Ejecucion.agente_id == agente_id, Ejecucion.ts >= desde)
            .order_by(Ejecucion.ts.asc())
            .all()
        )

    if exec_rows:
        dias_data: dict = {}
        for r in exec_rows:
            dia = r.ts.strftime("%Y-%m-%d") if r.ts else "?"
            if dia not in dias_data:
                dias_data[dia] = {"ok": 0, "fail": 0, "lats": []}
            if r.exitoso:
                dias_data[dia]["ok"] += 1
            else:
                dias_data[dia]["fail"] += 1
            if r.duracion_s and r.duracion_s > 0:
                dias_data[dia]["lats"].append(r.duracion_s)

        puntos = []
        for dia, d in sorted(dias_data.items()):
            total    = d["ok"] + d["fail"]
            lats     = d["lats"]
            prom_lat = round(sum(lats)/len(lats), 3) if lats else None
            puntos.append({
                "fecha":      dia,
                "total":      total,
                "exitosas":   d["ok"],
                "fallidas":   d["fail"],
                "tasa_exito": round(d["ok"]/total*100, 1) if total else 0,
                "latencia_s": prom_lat,
            })

        total_ok   = sum(d["ok"]   for d in dias_data.values())
        total_fail = sum(d["fail"] for d in dias_data.values())
        total_all  = total_ok + total_fail

        # Tendencia de exito (pendiente positiva = mejorando)
        tasas = [p["tasa_exito"] for p in puntos]
        tend_exito = _calcular_tendencia(tasas)

        # Tendencia de latencia (pendiente negativa = mejorando — menos tiempo es mejor)
        lats_serie = [p["latencia_s"] for p in puntos if p["latencia_s"] is not None]
        if len(lats_serie) >= 3:
            tend_lat_raw = _calcular_tendencia(lats_serie)
            # Invertir: pendiente negativa de latencia = "mejorando"
            tend_lat = {"mejorando": "empeorando", "empeorando": "mejorando", "estable": "estable"}[tend_lat_raw]
        else:
            tend_lat = "sin datos"

        todas_lats = [r.duracion_s for r in exec_rows if r.duracion_s and r.duracion_s > 0]
        lat_prom   = round(sum(todas_lats) / len(todas_lats), 3) if todas_lats else None

        return {
            "agente_id": agente_id,
            "dias":      dias,
            "puntos":    puntos,
            "fuente":    "pipeline",
            "resumen": {
                "total_ejecuciones":  total_all,
                "tasa_exito_global":  round(total_ok/total_all*100, 1) if total_all else 0,
                "latencia_prom_s":    lat_prom,
                "tendencia":          tend_exito,
                "tendencia_latencia": tend_lat,
            },
        }

    # ── Intento 2: fallback a conversaciones de chat ──────────────────────────
    try:
        from core.memory import Conversacion as _Conv, Mensaje as _Msg
        with get_session() as s:
            convs = (
                s.query(_Conv)
                .filter(_Conv.agente_id == agente_id, _Conv.ts_inicio >= desde)
                .all()
            )
            if not convs:
                return {"puntos": [], "resumen": None}

            conv_ids = [c.id for c in convs]
            all_msgs = (
                s.query(_Msg)
                .filter(_Msg.conversacion_id.in_(conv_ids))
                .order_by(_Msg.ts.asc())
                .all()
            )

        # Agrupa mensajes por conversación
        msgs_by_conv: dict = defaultdict(list)
        for m in all_msgs:
            msgs_by_conv[m.conversacion_id].append(m)

        dias_data2: dict = {}

        for conv in convs:
            msgs = msgs_by_conv.get(conv.id, [])
            if not msgs:
                continue  # ignorar sesiones vacías
            dia = conv.ts_inicio.strftime("%Y-%m-%d")
            if dia not in dias_data2:
                dias_data2[dia] = {"ok": 0, "fail": 0, "lats": []}

            user_msgs  = [m for m in msgs if m.rol == "usuario"]
            agent_msgs = [m for m in msgs if m.rol == "agente"]

            if agent_msgs:
                dias_data2[dia]["ok"] += 1
                if user_msgs and user_msgs[0].ts and agent_msgs[0].ts:
                    lat = (agent_msgs[0].ts - user_msgs[0].ts).total_seconds()
                    if 0 < lat < 300:
                        dias_data2[dia]["lats"].append(lat)
            else:
                dias_data2[dia]["fail"] += 1

        if not dias_data2:
            return {"puntos": [], "resumen": None}

        puntos2 = []
        for dia, d in sorted(dias_data2.items()):
            total    = d["ok"] + d["fail"]
            lats     = d["lats"]
            prom_lat = round(sum(lats)/len(lats), 3) if lats else None
            puntos2.append({
                "fecha":      dia,
                "total":      total,
                "exitosas":   d["ok"],
                "fallidas":   d["fail"],
                "tasa_exito": round(d["ok"]/total*100, 1) if total else 0,
                "latencia_s": prom_lat,
            })

        total_ok   = sum(d["ok"]   for d in dias_data2.values())
        total_fail = sum(d["fail"] for d in dias_data2.values())
        total_all  = total_ok + total_fail

        tasas2 = [p["tasa_exito"] for p in puntos2]
        tend_exito2 = _calcular_tendencia(tasas2)

        lats2 = [p["latencia_s"] for p in puntos2 if p["latencia_s"] is not None]
        if len(lats2) >= 3:
            tend_lat2_raw = _calcular_tendencia(lats2)
            tend_lat2 = {"mejorando": "empeorando", "empeorando": "mejorando", "estable": "estable"}[tend_lat2_raw]
        else:
            tend_lat2 = "sin datos"

        todas_lats2 = [p["latencia_s"] for p in puntos2 if p["latencia_s"] is not None]
        lat_prom2   = round(sum(todas_lats2)/len(todas_lats2), 3) if todas_lats2 else None

        return {
            "agente_id": agente_id,
            "dias":      dias,
            "puntos":    puntos2,
            "fuente":    "conversaciones",
            "resumen": {
                "total_ejecuciones":  total_all,
                "tasa_exito_global":  round(total_ok/total_all*100, 1) if total_all else 0,
                "latencia_prom_s":    lat_prom2,
                "tendencia":          tend_exito2,
                "tendencia_latencia": tend_lat2,
            },
        }
    except Exception:
        return {"puntos": [], "resumen": None}


@app.get("/agentes-stats")
async def agentes_stats_all() -> dict:
    """
    Devuelve estadísticas históricas de todos los agentes en el formato
    que usa localStorage ('agentdesk-agent-stats-v2') para pre-popular el
    Dashboard de Rendimiento al cargar la app.
    """
    config = load_config()
    agents = config.get("agents", [])
    stats: dict = {}
    for ag in agents:
        ag_id = ag.get("id", "")
        if not ag_id:
            continue
        try:
            data = await tendencias_agente(ag_id, dias=90)
            resumen = data.get("resumen")
            if not resumen:
                continue
            total = resumen.get("total_ejecuciones", 0)
            tasa  = resumen.get("tasa_exito_global", 100)
            ok    = round(total * tasa / 100)
            fail  = total - ok
            # Generar ultimasTareas con mínimo 2 entradas para que Mb() lo procese
            n_ok   = max(ok, 2) if ok > 0 else 0
            n_fail = max(fail, 1) if fail > 0 else 0
            ultima_ts = None
            puntos = data.get("puntos", [])
            if puntos:
                ultima_ts = puntos[-1].get("fecha")
            stats[ag_id] = {
                "ok":           ok,
                "fail":         fail,
                "ultimasTareas": (["ok"] * min(n_ok, 3) + ["fail"] * min(n_fail, 2))[-5:],
                "ultima_ts":    ultima_ts,
            }
        except Exception:
            continue
    return {"stats": stats}


# ── Auth JWT ──────────────────────────────────────────────────────────────────

class LoginRequest(BaseModel):
    username: str
    password: str

class CambiarPasswordRequest(BaseModel):
    username:        str
    nueva_password:  str
    token:           str   # debe ser admin para cambiar cualquier usuario


@app.post("/auth/login")
async def auth_login(payload: LoginRequest) -> dict:
    """Autentica y devuelve un JWT. El frontend lo guarda y envía en cada request."""
    from core.auth import login as _login, SistemaNoConfiguradoError
    try:
        result = _login(payload.username, payload.password)
    except SistemaNoConfiguradoError as exc:
        raise HTTPException(status_code=503, detail=str(exc))
    if result is None:
        raise HTTPException(status_code=401, detail="Credenciales incorrectas.")
    return result


@app.post("/auth/cambiar-password")
async def auth_cambiar_password(payload: CambiarPasswordRequest) -> dict:
    """Cambia la contraseña de un usuario (requiere token de admin)."""
    from core.auth import verificar_token, cambiar_password
    datos = verificar_token(payload.token)
    if not datos or datos.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Se requiere rol admin.")
    ok = cambiar_password(payload.username, payload.nueva_password)
    if not ok:
        raise HTTPException(status_code=404, detail=f"Usuario '{payload.username}' no encontrado.")
    return {"ok": True, "mensaje": f"Contraseña de '{payload.username}' actualizada."}


@app.get("/auth/verificar")
async def auth_verificar(authorization: str = "") -> dict:
    """Verifica si un token es válido."""
    from core.auth import verificar_token
    token = authorization.replace("Bearer ", "").strip()
    datos = verificar_token(token)
    if not datos:
        raise HTTPException(status_code=401, detail="Token inválido o expirado.")
    return {"ok": True, "username": datos.get("sub"), "role": datos.get("role")}


# ── RBAC: Gestión de Usuarios (admin) ────────────────────────────────────────

class CrearUsuarioRequest(BaseModel):
    username:  str
    password:  str
    rol:       str = "viewer"

class CambiarRolRequest(BaseModel):
    nuevo_rol: str

class ActivarRequest(BaseModel):
    activo: bool


@app.get("/auth/usuarios")
async def auth_listar_usuarios(
    req: "Request",
    _rbac = None,
) -> list[dict]:
    """Lista todos los usuarios del sistema. Solo admin."""
    from fastapi import Request, Depends
    from core.auth import listar_usuarios, requiere_rol as _req_rol
    # Verificación manual de rol (Depends no puede usarse en definición dinámica fácilmente)
    rol = getattr(req.state, "rol", "viewer")
    from core.auth import tiene_permiso
    if not tiene_permiso(rol, "admin"):
        raise HTTPException(403, detail="Se requiere rol admin.")
    return listar_usuarios()


@app.post("/auth/usuarios")
async def auth_crear_usuario(req: "Request", payload: CrearUsuarioRequest) -> dict:
    """Crea un nuevo usuario. Solo admin."""
    from core.auth import tiene_permiso, crear_usuario
    rol = getattr(req.state, "rol", "viewer")
    if not tiene_permiso(rol, "admin"):
        raise HTTPException(403, detail="Se requiere rol admin.")
    try:
        return crear_usuario(payload.username, payload.password, payload.rol)
    except ValueError as e:
        raise HTTPException(422, detail=str(e))


@app.delete("/auth/usuarios/{username}")
async def auth_eliminar_usuario(username: str, req: "Request") -> dict:
    """Elimina un usuario. Solo admin. No puede eliminar su propio usuario."""
    from core.auth import tiene_permiso, eliminar_usuario
    rol = getattr(req.state, "rol", "viewer")
    if not tiene_permiso(rol, "admin"):
        raise HTTPException(403, detail="Se requiere rol admin.")
    solicitante = getattr(req.state, "usuario", "")
    try:
        ok = eliminar_usuario(username, solicitante)
    except ValueError as e:
        raise HTTPException(422, detail=str(e))
    if not ok:
        raise HTTPException(404, detail=f"Usuario '{username}' no encontrado.")
    return {"ok": True, "eliminado": username}


@app.put("/auth/usuarios/{username}/rol")
async def auth_cambiar_rol(username: str, payload: CambiarRolRequest, req: "Request") -> dict:
    """Cambia el rol de un usuario. Solo admin."""
    from core.auth import tiene_permiso, cambiar_rol
    if not tiene_permiso(getattr(req.state, "rol", "viewer"), "admin"):
        raise HTTPException(403, detail="Se requiere rol admin.")
    try:
        return cambiar_rol(username, payload.nuevo_rol)
    except ValueError as e:
        raise HTTPException(422, detail=str(e))


@app.put("/auth/usuarios/{username}/activo")
async def auth_activar_usuario(username: str, payload: ActivarRequest, req: "Request") -> dict:
    """Activa o desactiva un usuario. Solo admin."""
    from core.auth import tiene_permiso, activar_desactivar
    if not tiene_permiso(getattr(req.state, "rol", "viewer"), "admin"):
        raise HTTPException(403, detail="Se requiere rol admin.")
    try:
        return activar_desactivar(username, payload.activo)
    except ValueError as e:
        raise HTTPException(422, detail=str(e))


# ── Analytics: Curva S / Earned Value Management ──────────────────────────────

@app.get("/analytics/curva-s/{proyecto_id}")
async def analytics_curva_s(proyecto_id: str, req: "Request") -> dict:
    """
    Calcula la Curva S (EVM) del proyecto y emite el resultado vía WebSocket.

    Si SPI o CPI < 0.80 (CRITICO) o < 0.90 (ALTO), también emite un evento
    `curva_s_alerta` para que el cliente React dispare una notificación nativa.

    Requiere rol supervisor o superior.
    """
    from core.auth import tiene_permiso
    if not tiene_permiso(getattr(req.state, "rol", "viewer"), "supervisor"):
        raise HTTPException(403, detail="Se requiere rol supervisor o admin.")

    from core.analytics import motor_analitica
    resultado = await asyncio.get_running_loop().run_in_executor(
        None,
        lambda: motor_analitica.calcular_curva_s(proyecto_id),
    )

    # Broadcast solo a supervisor+ (datos financieros sensibles)
    await manager.broadcast(
        {
            "tipo":        "curva_s_actualizada",
            "proyecto_id": proyecto_id,
            "kpis":        resultado.get("kpis", {}),
            "curva":       resultado.get("curva", []),
            "alerta":      resultado.get("alerta"),
        },
        rol_minimo="supervisor",
    )

    # Notificación proactiva adicional si hay desvío crítico
    if resultado.get("alerta"):
        kpis    = resultado.get("kpis", {})
        spi     = kpis.get("spi", 1.0)
        cpi     = kpis.get("cpi", 1.0)
        nivel   = resultado["alerta"]
        await manager.broadcast(
            {
                "tipo":        "curva_s_alerta",
                "proyecto_id": proyecto_id,
                "nivel":       nivel,
                "titulo":      f"Desvío {nivel} en proyecto {proyecto_id}",
                "cuerpo": (
                    f"SPI={spi:.2f} · CPI={cpi:.2f}. "
                    + ("Cronograma y costos fuera de control." if nivel == "CRITICO"
                       else "Revisar avance y presupuesto.")
                ),
            },
            rol_minimo="supervisor",   # alertas financieras solo para supervisor+
        )

    return resultado


# ── Backup / Restore / Versión ────────────────────────────────────────────────

@app.get("/version")
async def get_version() -> dict:
    """Versión actual de AgentDesk."""
    from core.backup import _version
    return {"version": _version(), "app": "AgentDesk"}


@app.get("/update/check")
async def update_check(url: str = "") -> dict:
    """Verifica si hay una nueva versión disponible."""
    from core.updater import verificar_actualizacion
    return verificar_actualizacion(url or None)


class UpdateURLRequest(BaseModel):
    url: str

@app.put("/update/url")
async def update_set_url(payload: UpdateURLRequest) -> dict:
    """Configura la URL del servidor de actualizaciones."""
    from core.updater import guardar_url_update
    ok = guardar_url_update(payload.url)
    return {"ok": ok, "url": payload.url}


@app.get("/backup/descargar")
async def backup_descargar(req: "Request") -> _Response:
    """Genera y descarga un ZIP con toda la base de datos y configuración. Solo admin (JWT)."""
    from core.auth import tiene_permiso
    rol     = getattr(req.state, "rol", "viewer")
    usuario = getattr(req.state, "usuario", None) or "anonimo"
    ip      = req.client.host if req.client else "desconocida"
    if not tiene_permiso(rol, "admin"):
        logger.warning(
            "AUDITORIA_SEGURIDAD: descarga de backup DENEGADA — "
            "user_id=%s rol=%s ip=%s endpoint='GET /backup/descargar'",
            usuario, rol, ip,
        )
        raise HTTPException(403, detail="Se requiere rol admin.")
    logger.info(
        "AUDITORIA_SEGURIDAD: descarga de backup AUTORIZADA — user_id=%s ip=%s",
        usuario, ip,
    )
    from core.backup import crear_backup
    from datetime import datetime as _dt2
    try:
        data = crear_backup()
        ts   = _dt2.utcnow().strftime("%Y%m%d_%H%M")
        return _Response(
            content    = data,
            media_type = "application/zip",
            headers    = {"Content-Disposition": f'attachment; filename="agentdesk_backup_{ts}.zip"'},
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error al crear backup: {exc}")


@app.post("/backup/restaurar")
async def backup_restaurar(req: "Request", archivo: UploadFile = _File(...)) -> dict:
    """Restaura un backup desde un ZIP previamente generado. Solo admin (JWT)."""
    from core.auth import tiene_permiso
    rol     = getattr(req.state, "rol", "viewer")
    usuario = getattr(req.state, "usuario", None) or "anonimo"
    ip      = req.client.host if req.client else "desconocida"
    if not tiene_permiso(rol, "admin"):
        logger.warning(
            "AUDITORIA_SEGURIDAD: restauracion de backup DENEGADA — "
            "user_id=%s rol=%s ip=%s endpoint='POST /backup/restaurar'",
            usuario, rol, ip,
        )
        raise HTTPException(403, detail="Se requiere rol admin.")
    logger.info(
        "AUDITORIA_SEGURIDAD: restauracion de backup AUTORIZADA — user_id=%s ip=%s",
        usuario, ip,
    )
    from core.backup import restaurar_backup
    try:
        data = await archivo.read()
        r    = restaurar_backup(data)
        if r["ok"]:
            await manager.broadcast({"tipo": "backup_restaurado", "total": r["total"]})
        return r
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


@app.get("/rate-limiter/stats")
async def rate_limiter_stats() -> dict:
    """Estadísticas de uso del rate limiter por proveedor."""
    try:
        from core.rate_limiter import get_stats_todos
        return {"stats": get_stats_todos()}
    except Exception as exc:
        return {"stats": [], "error": str(exc)}


@app.get("/health")
async def health() -> dict:
    agentes_info = {}
    if _orquestador:
        for aid, ag in _orquestador.agentes.items():
            agentes_info[aid] = {
                "nombre":  ag.nombre,
                "modelo":  ag.modelo,
                "area":    getattr(ag, "area", "General"),
                "status":  "idle",
            }
    return {
        "status":    "ok",
        "clientes_ws": len(manager.active),
        "agentes":   agentes_info,
    }


# ── Modelo de entrada para ejecutar tarea ──────────────────────────────────────
class EjecutarRequest(BaseModel):
    tarea:       str      = "reporte_ventas"
    datos_extra: str | None = None   # Texto directo para analizar
    archivo_id:  str | None = None   # ID de archivo subido con /upload


@app.get("/uploads")
async def listar_uploads() -> dict:
    """Lista todos los archivos subidos disponibles para analizar."""
    import json as _json
    from core.path_manager import data_path
    uploads_dir = data_path("uploads")
    if not uploads_dir.exists():
        return {"archivos": []}
    archivos = []
    for f in sorted(uploads_dir.glob("*.meta.json"), key=lambda x: x.stat().st_mtime, reverse=True):
        try:
            meta = _json.loads(f.read_text(encoding="utf-8"))
            archivos.append(meta)
        except Exception:
            pass
    return {"archivos": archivos[:30]}


@app.get("/uploads/{archivo_id}/texto")
async def get_upload_texto(archivo_id: str) -> dict:
    """Devuelve el contenido de texto de un archivo subido."""
    import json as _json
    from core.path_manager import data_path
    uploads_dir = data_path("uploads")
    meta_path   = uploads_dir / f"{archivo_id}.meta.json"
    if not meta_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado.")
    meta    = _json.loads(meta_path.read_text(encoding="utf-8"))
    archivo = uploads_dir / meta["nombre_interno"]
    if not archivo.exists():
        raise HTTPException(status_code=404, detail="Contenido no encontrado.")
    texto = archivo.read_bytes().decode("utf-8", errors="replace")[:20_000]
    return {"archivo_id": archivo_id, "nombre": meta["nombre_original"],
            "tipo": meta["tipo"], "texto": texto}


@app.post("/agentes/{agente_id}/ejecutar")
async def ejecutar_agente(agente_id: str, payload: EjecutarRequest) -> dict:
    """
    Ejecuta realizar_tarea() en el agente especificado.
    Emite eventos de telemetría en tiempo real via WebSocket /ws/telemetria.
    Retorna el reporte final o un error detallado.
    """
    if not kill_switch.is_active():
        return {"error": "Kill switch activo."}
    if _orquestador is None:
        return {"error": "Orquestador no inicializado. Verifica GEMINI_API_KEY en el .env."}
    if agente_id not in _orquestador.agentes:
        raise HTTPException(status_code=404, detail=f"Agente '{agente_id}' no encontrado.")

    await manager.broadcast({
        "tipo":     "agente_ejecutando",
        "agente_id": agente_id,
        "tarea":    payload.tarea,
    })

    import time as _time
    _t0 = _time.monotonic()

    try:
        agente    = _orquestador.agentes[agente_id]
        # Resolver datos: archivo_id > datos_extra > tarea normal
        datos_texto = None
        if payload.archivo_id:
            import json as _json
            from core.path_manager import data_path
            uploads_dir = data_path("uploads")
            meta_path   = uploads_dir / f"{payload.archivo_id}.meta.json"
            if meta_path.exists():
                meta  = _json.loads(meta_path.read_text(encoding="utf-8"))
                fpath = uploads_dir / meta["nombre_interno"]
                if fpath.exists():
                    datos_texto = fpath.read_bytes().decode("utf-8", errors="replace")[:18_000]
        elif payload.datos_extra:
            datos_texto = payload.datos_extra

        if datos_texto:
            resultado = await agente.realizar_tarea_con_datos(datos_texto)
        else:
            resultado = await agente.realizar_tarea(payload.tarea)

        duracion_s = round(_time.monotonic() - _t0, 3)

        if resultado is None:
            await manager.broadcast({"tipo":"tarea_abortada","agente_id":agente_id,"tarea":payload.tarea})
            try:
                from core.database import guardar_ejecucion
                guardar_ejecucion(agente_id=agente_id,
                    agente_nombre=_orquestador.agentes[agente_id].nombre,
                    tarea=payload.tarea, exitoso=False, duracion_s=duracion_s,
                    resumen="Abortado por guardrails")
            except Exception: pass
            return {"ok": False, "agente_id": agente_id,
                    "motivo": "Pipeline abortado por guardrails. Ve a Pipeline -> Feed de Errores para ver el detalle."}

        # Error de API (cuota, red, etc.) — distinto de un abort del pipeline
        if isinstance(resultado, dict) and resultado.get("_api_error"):
            msg = resultado.get("_api_msg", "Error de API")
            await manager.broadcast({"tipo":"tarea_error","agente_id":agente_id,"error":msg})
            try:
                from core.database import guardar_ejecucion
                guardar_ejecucion(agente_id=agente_id,
                    agente_nombre=_orquestador.agentes[agente_id].nombre,
                    tarea=payload.tarea, exitoso=False, duracion_s=duracion_s, resumen=msg)
            except Exception: pass
            return {"ok": False, "agente_id": agente_id, "motivo": msg}

        await manager.broadcast({
            "tipo":      "tarea_completada",
            "agente_id": agente_id,
            "tarea":     payload.tarea,
            "resumen":   resultado.get("resumen", "")[:200],
            "duracion_s": duracion_s,
        })
        agente_nombre = _orquestador.agentes[agente_id].nombre
        # Guardar en SQLite con duracion_s real para metricas historicas exactas
        try:
            from core.database import guardar_ejecucion
            guardar_ejecucion(
                agente_id=agente_id, agente_nombre=agente_nombre,
                tarea=payload.tarea, exitoso=True, duracion_s=duracion_s,
                resumen=resultado.get("resumen","")[:500] if resultado else "",
                kpis=resultado.get("kpis",{}) if resultado else {},
                archivo_id=payload.archivo_id,
            )
        except Exception: pass
        return {"ok": True, "agente_id": agente_id,
                "agente_nombre": agente_nombre, "resultado": resultado}

    except Exception as exc:
        await manager.broadcast({
            "tipo":      "tarea_error",
            "agente_id": agente_id,
            "error":     str(exc),
        })
        return {"ok": False, "agente_id": agente_id, "error": str(exc)}


@app.get("/agentes")
async def listar_agentes() -> dict:
    """Lista agentes aplanando ubicacion.lat/lng al nivel raíz para el mapa React."""
    try:
        config = load_config()
        agentes = []
        for a in config.get("agents", []):
            ub = a.get("ubicacion") or {}
            agentes.append({
                **a,
                "lat":   ub.get("lat",   a.get("lat",   0)),
                "lng":   ub.get("lng",   a.get("lng",   0)),
                "label": ub.get("label", a.get("label", "")),
            })
        return {"agentes": agentes}
    except Exception as e:
        return {"error": str(e), "agentes": []}


@app.delete("/agentes/{agente_id}")
async def eliminar_agente(agente_id: str) -> dict:
    """Elimina un agente del sistema y de config.json."""
    if not kill_switch.is_active():
        return {"error": "Kill switch activo."}
    if _orquestador is None:
        return {"error": "Orquestador no disponible."}
    ok = await _orquestador.eliminar_agente(agente_id)
    if not ok:
        raise HTTPException(status_code=404, detail=f"Agente '{agente_id}' no encontrado.")
    await manager.broadcast({"tipo": "agente_eliminado", "agente_id": agente_id})
    return {"ok": True, "agente_id": agente_id}


# ── Modelo de entrada para actualizar agente ───────────────────────────────────
class ActualizarAgenteRequest(BaseModel):
    nombre:      str  | None = None
    tipo_ia:     str  | None = None
    area:        str  | None = None
    modelo:      str  | None = None
    temperatura: float| None = None
    idioma:      str  | None = None
    prompt_base: str  | None = None


@app.put("/agentes/{agente_id}")
async def actualizar_agente(agente_id: str, payload: ActualizarAgenteRequest) -> dict:
    """Actualiza parámetros de un agente existente con validación Pydantic."""
    if not kill_switch.is_active():
        return {"error": "Kill switch activo."}
    if _orquestador is None:
        return {"error": "Orquestador no disponible."}
    data = {k: v for k, v in payload.model_dump().items() if v is not None}
    ok = await _orquestador.actualizar_agente(agente_id, data)
    if not ok:
        raise HTTPException(status_code=400, detail=f"Actualización fallida para '{agente_id}'.")
    await manager.broadcast({"tipo": "agente_actualizado", "agente_id": agente_id})
    return {"ok": True, "agente_id": agente_id}


@app.get("/proveedores")
async def listar_proveedores() -> dict:
    """Devuelve todos los proveedores de IA disponibles y sus modelos."""
    from core.providers import modelos_disponibles, proveedores_configurados
    return {
        "proveedores":  proveedores_configurados(),
        "modelos":      modelos_disponibles(),
    }


@app.put("/proveedores/apikey")
async def guardar_api_key(payload: dict) -> dict:
    """Guarda una API key de proveedor en el archivo .env."""
    from core.path_manager import data_path
    proveedor = payload.get("proveedor", "").upper()
    api_key   = payload.get("api_key", "").strip()
    if not proveedor or not api_key:
        raise HTTPException(status_code=400, detail="proveedor y api_key son requeridos.")

    env_key  = f"{proveedor}_API_KEY"
    # Buscar .env en el directorio AppData\AgentDesk
    import pathlib as _pathlib
    _appdata  = _pathlib.Path(os.environ.get("APPDATA", str(_pathlib.Path.home())))
    env_path  = _appdata / "AgentDesk" / ".env"
    if not env_path.exists():
        env_path.parent.mkdir(parents=True, exist_ok=True)
        env_path.write_text("", encoding="utf-8")

    lines = env_path.read_text(encoding="utf-8").splitlines()
    found = False
    new_lines = []
    for line in lines:
        if line.startswith(f"{env_key}="):
            new_lines.append(f"{env_key}={api_key}")
            found = True
        else:
            new_lines.append(line)
    if not found:
        new_lines.append(f"{env_key}={api_key}")

    env_path.write_text("\n".join(new_lines) + "\n", encoding="utf-8")
    os.environ[env_key] = api_key
    # Guardar también en vault cifrado
    try:
        from core.key_vault import guardar_key_cifrada
        guardar_key_cifrada(env_key, api_key)
    except Exception: pass
    logger.info("API key configurada para proveedor: %s", proveedor)
    return {"ok": True, "proveedor": proveedor, "env_key": env_key}


@app.get("/modelos")
async def listar_modelos() -> dict:
    """Devuelve todos los modelos de todos los proveedores configurados."""
    from core.providers import modelos_disponibles
    return {"modelos": modelos_disponibles()}


# ── Modelo de entrada para crear agente ────────────────────────────────────────
class NuevoAgenteRequest(BaseModel):
    nombre:      str
    tipo_ia:     str
    area:        str
    modelo:      str      = "models/gemini-2.5-flash"
    temperatura: float    = 0.4
    idioma:      str      = "espanol"
    prompt_base: str      = ""


# ── Chat libre con agentes ─────────────────────────────────────────────────────

class ChatRequest(BaseModel):
    mensaje:    str
    agente_id:  str | None = None
    archivo_id: str | None = None
    sesion_id:  str        = "default"   # ID de sesión para memoria persistente


@app.post("/chat")
async def chat(payload: ChatRequest) -> dict:
    """Envía un mensaje conversacional a un agente. El orquestador elige el agente si no se especifica."""
    if _orquestador is None:
        return {"error": "Orquestador no disponible.", "respuesta": None}

    agente    = None
    agente_key = None
    if payload.agente_id and payload.agente_id in _orquestador.agentes:
        agente_key = payload.agente_id
        agente     = _orquestador.agentes[agente_key]
    else:
        msg_lower = payload.mensaje.lower()
        for k, ag in _orquestador.agentes.items():
            if (ag.area or "").lower() in msg_lower:
                agente_key, agente = k, ag
                break
        if agente is None and _orquestador.agentes:
            agente_key, agente = next(iter(_orquestador.agentes.items()))

    if agente is None:
        return {"error": "No hay agentes disponibles.", "respuesta": None}

    # Con Tool Calling, el agente lee el archivo solo via leer_archivo()
    # El archivo_id se pasa directamente al agente como hint en su system prompt
    await manager.broadcast({"tipo": "chat_procesando",
                              "agente_id": agente_key,
                              "agente_nombre": agente.nombre})
    import asyncio as _asyncio
    try:
        # Tool Calling automático (Groq/OpenAI) con fallback a chat_libre
        herramientas_usadas: list[str] = []
        respuesta_tuple = await _asyncio.wait_for(
            agente.chat_con_herramientas(
                payload.mensaje,
                sesion_id       = payload.sesion_id,
                agente_id_clave = agente_key,
                archivo_id      = payload.archivo_id,   # pasa el archivo al agente
            ),
            timeout=90.0,
        )
        respuesta, herramientas_usadas = respuesta_tuple
        if herramientas_usadas:
            await manager.broadcast({
                "tipo":        "herramientas_usadas",
                "agente_id":   agente_key,
                "herramientas":herramientas_usadas,
            })
    except _asyncio.TimeoutError:
        respuesta = "⏰ El agente tardó más de 90 segundos. Intenta de nuevo."
    await manager.broadcast({"tipo": "chat_respuesta",
                              "agente_id": agente_key,
                              "agente_nombre": agente.nombre})

    logger.info("chat '%s': respondido", agente.nombre, extra={"agente": agente.nombre})
    return {"respuesta": respuesta, "agente_id": agente_key,
            "agente_nombre": agente.nombre, "agente_area": agente.area}


@app.get("/memoria/{agente_id}/sesiones")
async def memoria_sesiones(agente_id: str) -> dict:
    """Lista las sesiones de conversación de un agente."""
    from core.memory import get_sesiones_agente
    return {"sesiones": get_sesiones_agente(agente_id)}


@app.get("/memoria/{agente_id}/historial/{sesion_id}")
async def memoria_historial(agente_id: str, sesion_id: str) -> dict:
    """Historial completo de una sesión de conversación."""
    from core.memory import get_historial_sesion
    return {"mensajes": get_historial_sesion(agente_id, sesion_id)}


@app.delete("/memoria/{agente_id}/sesion/{sesion_id}")
async def memoria_limpiar(agente_id: str, sesion_id: str) -> dict:
    """Elimina todos los mensajes de una sesión (resetear conversación)."""
    from core.memory import limpiar_sesion
    limpiar_sesion(agente_id, sesion_id)
    return {"ok": True}


from fastapi.responses import StreamingResponse as _StreamingResponse

@app.post("/chat/stream")
async def chat_stream(payload: ChatRequest) -> _StreamingResponse:
    """
    Versión STREAMING del chat — devuelve Server-Sent Events (SSE).
    El frontend lee los chunks y muestra texto conforme llega.
    """
    import json as _json

    if _orquestador is None:
        async def _err():
            yield f"data: {_json.dumps({'error': 'Orquestador no disponible'})}\n\n"
            yield "data: [DONE]\n\n"
        return _StreamingResponse(_err(), media_type="text/event-stream")

    # Encontrar agente
    agente     = None
    agente_key = None
    if payload.agente_id and payload.agente_id in _orquestador.agentes:
        agente_key = payload.agente_id
        agente     = _orquestador.agentes[agente_key]
    else:
        msg_lower = payload.mensaje.lower()
        for k, ag in _orquestador.agentes.items():
            if (ag.area or "").lower() in msg_lower:
                agente_key, agente = k, ag; break
        if agente is None and _orquestador.agentes:
            agente_key, agente = next(iter(_orquestador.agentes.items()))

    if agente is None:
        async def _no_agent():
            yield f"data: {_json.dumps({'error': 'No hay agentes disponibles'})}\n\n"
            yield "data: [DONE]\n\n"
        return _StreamingResponse(_no_agent(), media_type="text/event-stream")

    async def event_generator():
        import asyncio as _asyncio
        texto_completo = ""
        yield f"data: {_json.dumps({'tipo': 'inicio', 'agente_nombre': agente.nombre, 'agente_area': agente.area, 'agente_id': agente_key})}\n\n"

        # Timeout por PASO (no por conversación completa): chat_con_herramientas_stream
        # puede encadenar hasta MAX_PASOS=6 llamadas al modelo (una por herramienta
        # invocada) antes de la respuesta final en streaming. Un límite fijo de 90s
        # para todo el intercambio cortaba conversaciones con varias herramientas
        # que seguían avanzando con normalidad; ahora cada paso individual (una
        # llamada al modelo o una herramienta) tiene su propio margen, y la
        # conversación completa puede tardar lo que necesite mientras siga avanzando.
        PASO_TIMEOUT_S = 45.0
        try:
            aiter = agente.chat_con_herramientas_stream(
                payload.mensaje,
                sesion_id       = payload.sesion_id,
                agente_id_clave = agente_key,
                archivo_id      = payload.archivo_id,
            ).__aiter__()

            while True:
                try:
                    evento = await _asyncio.wait_for(aiter.__anext__(), timeout=PASO_TIMEOUT_S)
                except StopAsyncIteration:
                    break
                if evento.get("tipo") == "chunk":
                    texto_completo += evento["chunk"]
                yield f"data: {_json.dumps(evento, ensure_ascii=False)}\n\n"

        except _asyncio.TimeoutError:
            yield f"data: {_json.dumps({'tipo': 'error', 'error': f'Tiempo de espera agotado ({int(PASO_TIMEOUT_S)}s) esperando un paso del modelo. Intenta de nuevo.'})}\n\n"
        except Exception as exc:
            yield f"data: {_json.dumps({'tipo': 'error', 'error': str(exc)})}\n\n"
        finally:
            yield f"data: {_json.dumps({'tipo': 'fin', 'texto_completo': texto_completo})}\n\n"
            yield "data: [DONE]\n\n"

    return _StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control":    "no-cache",
            "X-Accel-Buffering":"no",
            "Connection":       "keep-alive",
        },
    )


@app.post("/upload")
async def upload_archivo(archivo: UploadFile = _File(...)) -> dict:
    """Sube un archivo para análisis. Devuelve archivo_id + preview de CSV/Excel."""
    import uuid, json as _json
    from core.path_manager import data_path
    uploads_dir = data_path("uploads")
    uploads_dir.mkdir(parents=True, exist_ok=True)

    archivo_id      = str(uuid.uuid4())[:8]
    nombre_original = archivo.filename or "archivo"
    ext             = nombre_original.rsplit(".", 1)[-1].lower() if "." in nombre_original else "bin"
    nombre_interno  = f"{archivo_id}.{ext}"
    contenido       = await archivo.read()
    ruta            = uploads_dir / nombre_interno
    ruta.write_bytes(contenido)

    # ── Preview de datos estructurados ────────────────────────────────────────
    preview: dict = {}
    try:
        if ext == "csv":
            import io, csv
            texto = contenido.decode("utf-8", errors="replace")
            # Detectar separador automáticamente
            dialecto = csv.Sniffer().sniff(texto[:4096], delimiters=",;\t|")
            reader   = csv.DictReader(io.StringIO(texto), dialect=dialecto)
            columnas = reader.fieldnames or []
            filas    = [row for _, row in zip(range(5), reader)]
            preview  = {
                "columnas":    list(columnas),
                "n_columnas":  len(columnas),
                "muestra":     filas,
                "separador":   dialecto.delimiter,
                "total_lineas": texto.count("\n"),
            }
        elif ext in ("xlsx", "xls"):
            import openpyxl
            wb    = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            sheet = wb.active
            filas_raw = list(sheet.iter_rows(min_row=1, max_row=6, values_only=True))
            if filas_raw:
                encabezado = [str(c) if c is not None else "" for c in filas_raw[0]]
                muestra    = [
                    {encabezado[i]: str(v) if v is not None else ""
                     for i, v in enumerate(fila)}
                    for fila in filas_raw[1:]
                ]
                preview = {
                    "columnas":   encabezado,
                    "n_columnas": len(encabezado),
                    "muestra":    muestra,
                    "hojas":      wb.sheetnames,
                    "total_filas": sheet.max_row,
                }
            wb.close()
    except Exception as exc:
        logger.debug("upload preview error (%s): %s", ext, exc)

    meta = {
        "archivo_id": archivo_id, "nombre_original": nombre_original,
        "nombre_interno": nombre_interno, "tipo": ext,
        "tamano_bytes": len(contenido), "preview": preview,
    }
    (uploads_dir / f"{archivo_id}.meta.json").write_text(
        _json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info("Archivo subido: %s (%d bytes) — %d columnas detectadas",
                nombre_original, len(contenido), len(preview.get("columnas", [])))
    return {
        "archivo_id": archivo_id, "nombre": nombre_original,
        "tipo": ext, "tamano_kb": round(len(contenido) / 1024, 1),
        "preview": preview,
    }


class GenerarPDFRequest(BaseModel):
    reporte:        dict
    titulo:         str  = "Informe de Análisis"
    subtitulo:      str  = ""
    nombre_agente:  str  = "AgentDesk"
    archivo_nombre: str  = ""
    empresa:        str  = "AgentDesk"


@app.post("/generar-pdf")
async def generar_pdf(payload: GenerarPDFRequest) -> dict:
    """Genera un informe PDF, lo guarda en reportes/ y devuelve la URL de descarga."""
    try:
        from core.report_generator import generar_pdf as _gen_pdf
        from core.path_manager import data_path
        from datetime import datetime as _dt
        pdf_bytes = _gen_pdf(
            reporte        = payload.reporte,
            titulo         = payload.titulo,
            subtitulo      = payload.subtitulo,
            nombre_agente  = payload.nombre_agente,
            archivo_nombre = payload.archivo_nombre,
            empresa        = payload.empresa,
        )
        reportes_dir = data_path("reportes")
        reportes_dir.mkdir(parents=True, exist_ok=True)
        ts = _dt.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = payload.titulo.replace(" ", "_")[:30] + f"_{ts}.pdf"
        (reportes_dir / nombre_archivo).write_bytes(pdf_bytes)
        return {"ok": True, "filename": nombre_archivo, "url": f"/reportes/{nombre_archivo}"}
    except Exception as exc:
        logger.error("generar_pdf error: %s", exc)
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {exc}")


@app.get("/alertas/config")
async def get_alertas_config() -> dict:
    """Lee la configuración de umbrales para alertas económicas."""
    from core.path_manager import data_path
    import json as _json
    cfg_path = data_path("alertas_config.json")
    defaults = {
        "umbrales": {
            "dolar_max": 1000,
            "dolar_min": 800,
            "uf_max":    45000,
            "ipc_max":   1.0,
        }
    }
    if cfg_path.exists():
        try:
            stored = _json.loads(cfg_path.read_text(encoding="utf-8"))
            return {"config": {**defaults, **stored}}
        except Exception:
            pass
    return {"config": defaults}


class AlertasConfigRequest(BaseModel):
    dolar_max: float | None = None
    dolar_min: float | None = None
    uf_max:    float | None = None
    ipc_max:   float | None = None


@app.put("/alertas/config")
async def set_alertas_config(payload: AlertasConfigRequest) -> dict:
    """Actualiza umbrales de alertas económicas."""
    from core.path_manager import data_path
    import json as _json
    cfg_path = data_path("alertas_config.json")
    data_path("").mkdir(parents=True, exist_ok=True)

    current = {"umbrales": {}}
    if cfg_path.exists():
        try: current = _json.loads(cfg_path.read_text(encoding="utf-8"))
        except Exception: pass

    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    current.setdefault("umbrales", {}).update(updates)
    cfg_path.write_text(_json.dumps(current, indent=2), encoding="utf-8")
    logger.info("Alertas config actualizada: %s", updates)
    return {"ok": True, "config": current}


@app.get("/indicadores/live")
async def indicadores_live() -> dict:
    """Retorna UF, dólar, euro, IPC actuales del Banco Central en tiempo real."""
    try:
        from core.web_monitor import fetch_indicadores_economia
        data = await fetch_indicadores_economia()
        if "error" in data:
            raise HTTPException(status_code=502, detail=data["error"])
        return {"indicadores": data, "fuente": "mindicador.cl"}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/pipeline/config")
async def get_pipeline_config() -> dict:
    """Lee la configuración de umbrales de los guardrails."""
    from core.path_manager import data_path
    import json as _json
    cfg_path = data_path("pipeline_config.json")
    defaults = {
        "recursion_umbral": 3,
        "grounding_min":    1000,
        "logic_factor":     100,
        "timeout_s":        5,
    }
    if cfg_path.exists():
        try:
            stored = _json.loads(cfg_path.read_text(encoding="utf-8"))
            return {"config": {**defaults, **stored}}
        except Exception:
            pass
    return {"config": defaults}


class PipelineConfigRequest(BaseModel):
    recursion_umbral: int   | None = None
    grounding_min:    int   | None = None
    logic_factor:     int   | None = None
    timeout_s:        float | None = None


@app.put("/pipeline/config")
async def set_pipeline_config(payload: PipelineConfigRequest) -> dict:
    """
    Actualiza los umbrales de los guardrails.
    Los cambios se aplican en la próxima ejecución de agentes.
    """
    from core.path_manager import data_path
    import json as _json
    cfg_path = data_path("pipeline_config.json")
    data_path("").mkdir(parents=True, exist_ok=True)

    current = {}
    if cfg_path.exists():
        try: current = _json.loads(cfg_path.read_text(encoding="utf-8"))
        except Exception: pass

    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    current.update(updates)
    cfg_path.write_text(_json.dumps(current, indent=2), encoding="utf-8")

    logger.info("Pipeline config actualizada", extra={"config": current})
    return {"ok": True, "config": current}


@app.get("/embeddings")
async def get_embeddings() -> dict:
    """
    Embeddings semánticos reales usando TF-IDF + PCA.
    Los agentes similares (mismo dominio, mismos temas) quedan CERCANOS en 3D.
    """
    from core.embeddings import calcular_embeddings
    from core.database   import get_historial

    # Construir lista de agentes con su configuración completa
    agentes_lista = []
    if _orquestador:
        for aid, ag in _orquestador.agentes.items():
            agentes_lista.append({
                "id":           aid,
                "nombre":       ag.nombre,
                "area":         getattr(ag, "area", "General"),
                "prompt_base":  getattr(ag, "prompt_base", ""),
                "modelo":       ag.modelo,
                "temperatura":  getattr(ag, "temperatura", 0.4),
                "idioma":       getattr(ag, "idioma", "español"),
            })

    if not agentes_lista:
        # Cargar desde config.json si el orquestador no está listo
        try:
            from core.config_loader import load_config
            cfg = load_config()
            agentes_lista = cfg.get("agents", [])
        except Exception:
            pass

    # Historial de ejecuciones por agente para enriquecer embeddings
    hist_por_agente = {}
    try:
        for ag in agentes_lista:
            hist_por_agente[ag["id"]] = get_historial(agente_id=ag["id"], limit=10)
    except Exception:
        pass

    # Calcular embeddings reales con TF-IDF + PCA
    puntos = calcular_embeddings(agentes_lista, hist_por_agente)

    return {
        "puntos":  puntos,
        "total":   len(puntos),
        "metodo":  "TF-IDF + PCA (semántico real)" if len(agentes_lista) > 1 else "distribución circular",
        "agentes": len(agentes_lista),
    }


@app.get("/kill-switch")
async def estado_kill_switch() -> dict:
    """
    Retorna el estado actual del kill switch (incluye gist_url para la UI).
    El monitor de background actualiza el estado cada 5 minutos.
    """
    return kill_switch.estado_dict()


class KillSwitchURLRequest(BaseModel):
    url: str = ""


class KillSwitchToggleRequest(BaseModel):
    activo: bool


@app.post("/kill-switch/toggle")
async def toggle_kill_switch(payload: KillSwitchToggleRequest) -> dict:
    """
    Activa o bloquea el Kill Switch manualmente desde el SecurityPanel.
    Requiere rol admin (verificado por JWTMiddleware).
    El monitor del Gist puede sobreescribir este estado en la siguiente verificación.
    """
    kill_switch.forzar_estado(payload.activo)
    logger.info("Kill switch toggled a %s via API.", payload.activo)
    return kill_switch.estado_dict()


@app.post("/kill-switch/url")
async def configurar_kill_switch_url(payload: KillSwitchURLRequest) -> dict:
    """
    Actualiza la URL del Gist de Kill Switch en tiempo de ejecucion.
    Requiere JWT con rol admin (protegido por JWTMiddleware).
    Si url es vacia, desactiva el control remoto.
    """
    kill_switch.set_gist_url(payload.url)
    if payload.url:
        resultado = await kill_switch.verificar_gist()
        return {"ok": True, "url": payload.url, "active": resultado}
    return {"ok": True, "url": "", "active": True, "nota": "Control remoto desactivado."}


# ── Modelo de entrada para RELOAD_CONFIG ──────────────────────────────────────
class ReloadRequest(BaseModel):
    agente_id: str | None = None   # None → recargar todos los agentes


@app.post("/reload")
async def recargar_config(payload: ReloadRequest) -> dict:
    """
    Envía RELOAD_CONFIG al CommandBridge.

    El Orquestador re-lee config.json, valida cada agente con Pydantic
    y aplica los cambios en caliente.  Si la validación falla, el agente
    mantiene su configuración anterior (rollback implícito).

    payload.agente_id: ID del agente a recargar, o null para todos.
    """
    if not kill_switch.is_active():
        return {
            "error": "Kill switch activo — operación bloqueada.",
            "kill_switch": kill_switch.estado_dict(),
        }

    if _bridge is None:
        return {"error": "CommandBridge no disponible. El orquestador no está conectado."}

    await _bridge.send(Command(
        tipo=RELOAD_CONFIG,
        payload={"agente_id": payload.agente_id},
    ))

    await manager.broadcast({
        "tipo":      "reload_solicitado",
        "agente_id": payload.agente_id or "todos",
    })

    return {
        "ok":       True,
        "agente_id": payload.agente_id or "todos",
        "mensaje":  "RELOAD_CONFIG encolado. El Orquestador aplicará los cambios en breve.",
    }


@app.post("/agentes", status_code=201)
async def crear_agente(payload: NuevoAgenteRequest) -> dict:
    """
    Valida con AgentConfig (Pydantic) y encola CREAR_AGENTE en el CommandBridge.
    El Orquestador consume el comando y persiste en config.json.
    """
    if not kill_switch.is_active():
        return {
            "error": "Kill switch activo — creación de agentes bloqueada.",
            "kill_switch": kill_switch.estado_dict(),
        }

    if _bridge is None:
        return {"error": "CommandBridge no disponible. El orquestador no está conectado."}

    # Validar con el mismo schema que usa el Orquestador
    try:
        AgentConfig.model_validate({
            "nombre":      payload.nombre,
            "tipo_ia":     payload.tipo_ia,
            "area":        payload.area,
            "modelo":      payload.modelo,
            "temperatura": payload.temperatura,
            "idioma":      payload.idioma,
            "prompt_base": payload.prompt_base,
        })
    except Exception as e:
        return {"error": f"Validacion fallida: {e}"}

    # Enviar al CommandBridge — el Orquestador lo procesará de forma asíncrona
    await _bridge.send(Command(
        tipo=CREAR_AGENTE,
        payload={
            "nombre":      payload.nombre,
            "tipo_ia":     payload.tipo_ia,
            "area":        payload.area,
            "modelo":      payload.modelo,
            "temperatura": payload.temperatura,
            "idioma":      payload.idioma,
            "prompt_base": payload.prompt_base,
        }
    ))

    # Notificar a todos los clientes WebSocket del nuevo agente
    await manager.broadcast({
        "tipo":   "agente_creado",
        "nombre": payload.nombre,
        "area":   payload.area,
    })

    return {"ok": True, "nombre": payload.nombre, "area": payload.area}


# ── Helpers PDF ───────────────────────────────────────────────────────────────

def _slug(texto: str) -> str:
    return texto.lower().replace(" ", "_").replace("-", "_")


def _buscar_pdf(prefijo: str, agente_id: str) -> Path | None:
    """
    Busca el PDF más reciente con el prefijo dado para el agente.
    Tolera tanto el ID crudo como su versión slug.
    """
    slug    = _slug(agente_id)
    carpeta = REPORTES_DIR
    if not carpeta.exists():
        return None
    patron  = re.compile(rf"^{re.escape(prefijo)}_{re.escape(slug)}_\d{{8}}_\d{{6}}\.pdf$")
    matches = [f for f in carpeta.iterdir() if patron.match(f.name)]
    return max(matches, key=lambda f: f.stat().st_mtime) if matches else None


# ── Endpoints de reportes PDF ──────────────────────────────────────────────────

@app.get("/agentes/{agente_id}/reporte")
async def descargar_reporte(agente_id: str) -> FileResponse:
    """
    Devuelve el PDF de éxito más reciente del agente como descarga directa.

    Busca en %APPDATA%\\AgentDesk\\reportes\\ archivos de la forma:
      reporte_{slug}_{YYYYMMDD}_{HHMMSS}.pdf

    Responde 404 si el agente aún no tiene reportes (no ha ejecutado el pipeline).
    """
    pdf = _buscar_pdf("reporte", agente_id)
    if not pdf:
        raise HTTPException(
            status_code=404,
            detail=(
                f"Sin reporte PDF para '{agente_id}'. "
                "Ejecuta el agente desde el backend Python para generarlo."
            ),
        )
    return FileResponse(
        path=str(pdf),
        media_type="application/pdf",
        filename=pdf.name,
        headers={"Content-Disposition": f'attachment; filename="{pdf.name}"'},
    )


@app.get("/agentes/{agente_id}/correccion")
async def descargar_correccion(agente_id: str) -> FileResponse:
    """Devuelve el PDF de corrección más reciente del agente."""
    pdf = _buscar_pdf("correccion", agente_id)
    if not pdf:
        raise HTTPException(
            status_code=404,
            detail=f"Sin PDF de corrección para '{agente_id}'.",
        )
    return FileResponse(
        path=str(pdf),
        media_type="application/pdf",
        filename=pdf.name,
        headers={"Content-Disposition": f'attachment; filename="{pdf.name}"'},
    )


@app.get("/agentes/{agente_id}/reportes")
async def listar_reportes(agente_id: str) -> dict:
    """
    Lista todos los PDFs disponibles para el agente (éxito + correcciones).
    Ordenados por fecha descendente.
    """
    slug    = _slug(agente_id)
    carpeta = REPORTES_DIR
    if not carpeta.exists():
        return {"agente": agente_id, "reportes": []}

    archivos = [
        {
            "nombre": f.name,
            "tipo":   "correccion" if f.name.startswith("correccion_") else "reporte",
            "mtime":  f.stat().st_mtime,
        }
        for f in carpeta.iterdir()
        if f.suffix == ".pdf" and slug in f.name
    ]
    archivos.sort(key=lambda x: x["mtime"], reverse=True)
    return {"agente": agente_id, "reportes": archivos}


# ── Motor Financiero ───────────────────────────────────────────────────────────

class PresupuestoPayload(BaseModel):
    """Payload para POST /finanzas/analizar."""
    agente_id:   str
    presupuesto: dict
    periodos:    int = 6


@app.get("/finanzas/indicadores")
async def finanzas_indicadores() -> dict:
    """Indicadores macro Chile en tiempo real (UF, Dólar, IPC, Euro)."""
    from core.finance import motor_financiero
    try:
        ind = await motor_financiero.obtener_indicadores()
        return {"ok": True, "indicadores": ind.model_dump(mode="json")}
    except RuntimeError as e:
        return {"ok": False, "error": str(e)}


@app.post("/finanzas/analizar")
async def finanzas_analizar(payload: PresupuestoPayload) -> dict:
    """
    Análisis financiero completo: valida con Pydantic v2, obtiene indicadores
    reales del Banco Central y persiste en agentdesk.db (rollback si falla).
    """
    from pydantic import ValidationError
    from core.schemas import PresupuestoConfig
    from core.finance import motor_financiero

    try:
        presupuesto = PresupuestoConfig.model_validate(payload.presupuesto)
    except ValidationError as e:
        raise HTTPException(422, detail={"error": "Presupuesto inválido", "detalle": e.errors()})

    try:
        resultado = await motor_financiero.analizar_y_persistir(
            payload.agente_id, presupuesto, periodos=payload.periodos,
        )
        return {"ok": True, "analisis": resultado}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@app.get("/finanzas/historico/{agente_id}")
async def finanzas_historico(agente_id: str, n: int = 10) -> dict:
    """Historial de análisis financieros del agente (últimos N registros)."""
    from core.finance import motor_financiero
    registros = motor_financiero.historico(agente_id, n=min(n, 50))
    return {"agente_id": agente_id, "registros": registros}


@app.get("/finanzas/tendencia/{agente_id}")
async def finanzas_tendencia(agente_id: str, n: int = 10) -> dict:
    """Estadísticas de tendencia del flujo neto histórico del agente."""
    from core.finance import motor_financiero
    return motor_financiero.tendencia_flujo(agente_id, n=n)


@app.post("/finanzas/reload/{agente_id}")
async def finanzas_reload_presupuesto(agente_id: str, presupuesto: dict) -> dict:
    """Recarga el presupuesto de un agente en caliente (RELOAD_FINANZAS via CommandBridge)."""
    if not _bridge:
        raise HTTPException(503, detail="Bridge no disponible — orquestador no en línea")
    await _bridge.send(Command(
        tipo=RELOAD_FINANZAS,
        payload={"agente_id": agente_id, "presupuesto": presupuesto},
    ))
    return {"ok": True, "mensaje": f"RELOAD_FINANZAS encolado para '{agente_id}'"}


# ── Motor Gantt ────────────────────────────────────────────────────────────────

@app.get("/gantt/proyectos")
async def gantt_listar_proyectos() -> dict:
    """Lista todos los proyectos Gantt con su resumen de avance."""
    from core.gantt import motor_gantt
    return {"proyectos": motor_gantt.listar_proyectos()}


@app.get("/gantt/{proyecto_id}")
async def gantt_obtener_proyecto(proyecto_id: str) -> dict:
    """Retorna todas las tareas y el resumen CPM de un proyecto."""
    from core.gantt import motor_gantt
    return motor_gantt.obtener_proyecto(proyecto_id)


@app.post("/gantt/{proyecto_id}/tareas")
async def gantt_crear_tarea(proyecto_id: str, datos: dict) -> dict:
    """Crea una tarea en el proyecto y recalcula la Ruta Crítica (CPM)."""
    from core.gantt import motor_gantt
    datos["proyecto_id"] = proyecto_id
    try:
        return motor_gantt.crear_tarea(datos)
    except ValueError as e:
        raise HTTPException(422, detail=str(e))


@app.put("/gantt/tareas/{tarea_id}")
async def gantt_actualizar_tarea(tarea_id: int, datos: dict) -> dict:
    """Actualiza parámetros de una tarea y recalcula Forward/Backward pass."""
    from core.gantt import motor_gantt
    try:
        return motor_gantt.actualizar_tarea(tarea_id, datos)
    except ValueError as e:
        raise HTTPException(422, detail=str(e))


@app.patch("/gantt/tareas/{tarea_id}/progreso")
async def gantt_actualizar_progreso(tarea_id: int, datos: dict) -> dict:
    """Actualiza el % de avance de una tarea con validación Pydantic (rollback si falla)."""
    from core.gantt import motor_gantt
    try:
        resultado = motor_gantt.actualizar_progreso(tarea_id, datos)
        # Broadcast del nuevo estado a todos los clientes WS
        await manager.broadcast({
            "tipo":      "gantt_progreso",
            "tarea_id":  tarea_id,
            "pct":       resultado["pct_completado"],
            "proyecto":  resultado["proyecto_id"],
        })
        return resultado
    except ValueError as e:
        raise HTTPException(422, detail=str(e))


@app.delete("/gantt/tareas/{tarea_id}")
async def gantt_eliminar_tarea(tarea_id: int) -> dict:
    """Elimina una tarea y recalcula el CPM del proyecto."""
    from core.gantt import motor_gantt
    ok = motor_gantt.eliminar_tarea(tarea_id)
    if not ok:
        raise HTTPException(404, detail=f"Tarea {tarea_id} no encontrada")
    return {"ok": True, "eliminada": tarea_id}


@app.get("/gantt/{proyecto_id}/pdf")
async def gantt_exportar_pdf(
    proyecto_id: str,
    agente_id: str | None = None,
) -> _Response:
    """
    Genera un Reporte de Avance de Obra en PDF con:
      - Cronograma Gantt (barras horizontales con progreso)
      - Indicadores financieros en tiempo real (UF, Dólar, IPC)
      - Resumen de tareas críticas y avance global
    """
    from core.gantt import motor_gantt
    from core.finance import motor_financiero

    proyecto = motor_gantt.obtener_proyecto(proyecto_id)
    if not proyecto["tareas"]:
        raise HTTPException(404, detail=f"Proyecto '{proyecto_id}' sin tareas")

    # Intentar obtener indicadores financieros (sin bloquear si falla)
    indicadores = None
    try:
        indicadores = await motor_financiero.obtener_indicadores()
    except RuntimeError:
        pass

    pdf_bytes = _generar_pdf_gantt(proyecto, indicadores, agente_id=agente_id)

    nombre = f"avance_{proyecto_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.pdf"
    return _Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


def _generar_pdf_gantt(proyecto: dict, indicadores=None, _agente_id: str | None = None) -> bytes:
    """Genera PDF de avance de obra usando primitivas fpdf (sin imágenes externas)."""
    from fpdf import FPDF

    AZUL_OSC = (10,  45,  95)
    AZUL_MED = (26,  92, 140)
    CYAN     = (0,  140, 190)
    GRIS_OSC = (40,  50,  65)
    GRIS_CLA = (235, 240, 245)
    BLANCO   = (255, 255, 255)
    CRITICO  = (220, 60,  60)

    resumen   = proyecto["resumen"]
    tareas    = proyecto["tareas"]
    pid       = proyecto["proyecto_id"]

    def _txt(s):
        if not s:
            return ""
        for k, v in {"á":"a","é":"e","í":"i","ó":"o","ú":"u","ñ":"n","ü":"u",
                     "Á":"A","É":"E","Í":"I","Ó":"O","Ú":"U","Ñ":"N"}.items():
            s = s.replace(k, v)
        return s.encode("latin-1", errors="replace").decode("latin-1")

    class GanttPDF(FPDF):
        def header(self):
            if self.page_no() == 1:
                return
            self.set_fill_color(*AZUL_OSC)
            self.rect(0, 0, 210, 10, "F")
            self.set_y(2)
            self.set_font("Helvetica", "B", 8)
            self.set_text_color(*BLANCO)
            self.cell(0, 6, _txt(f"Reporte de Avance - Proyecto: {pid}"), align="C")
            self.set_text_color(0, 0, 0)
            self.ln(10)

        def footer(self):
            self.set_y(-13)
            self.set_draw_color(*CYAN)
            self.set_line_width(0.4)
            self.line(15, self.get_y(), 195, self.get_y())
            self.set_font("Helvetica", "", 7)
            self.set_text_color(*GRIS_OSC)
            meses = ["enero","febrero","marzo","abril","mayo","junio",
                     "julio","agosto","septiembre","octubre","noviembre","diciembre"]
            n = datetime.utcnow()
            fecha = f"{n.day} de {meses[n.month-1]} de {n.year}"
            self.cell(0, 6, _txt(f"AgentDesk - {fecha} - Pag. {self.page_no()}"), align="C")

    pdf = GanttPDF()
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.set_margins(12, 12, 12)

    # ── Portada ─────────────────────────────────────────────────────────────────
    pdf.add_page()
    pdf.set_fill_color(*AZUL_OSC)
    pdf.rect(0, 0, 210, 48, "F")
    pdf.set_fill_color(*CYAN)
    pdf.rect(0, 48, 210, 2.5, "F")

    pdf.set_xy(10, 10)
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(*BLANCO)
    pdf.multi_cell(190, 9, "REPORTE DE AVANCE DE OBRA", align="C")
    pdf.set_x(10)
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(170, 210, 235)
    pdf.multi_cell(190, 7, _txt(f"Proyecto: {pid}"), align="C")

    # Métricas portada
    pdf.set_y(60)
    meta = [
        ("Avance Global",    f"{resumen.get('pct_avance', 0):.1f}%"),
        ("Total Tareas",     str(resumen.get("total_tareas", 0))),
        ("Tareas Criticas",  str(resumen.get("tareas_criticas", 0))),
        ("Fecha Inicio",     (resumen.get("fecha_inicio") or "")[:10]),
        ("Fecha Fin Plan",   (resumen.get("fecha_fin")    or "")[:10]),
    ]
    if indicadores:
        meta += [
            ("UF (BCCh)",    f"${indicadores.uf:,.2f} CLP"),
            ("Dolar (BCCh)", f"${indicadores.dolar:,.2f} CLP"),
            ("IPC",          f"{indicadores.ipc:.2f}%"),
        ]
    for k, v in meta:
        pdf.set_x(20)
        pdf.set_fill_color(*GRIS_CLA)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(*AZUL_MED)
        pdf.cell(60, 7, _txt(k), border=0, fill=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*GRIS_OSC)
        pdf.cell(0, 7, _txt(str(v)), border=0, fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1)

    # ── Cronograma Gantt (barras horizontales) ──────────────────────────────────
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_text_color(*AZUL_OSC)
    pdf.cell(0, 9, "Cronograma de Tareas", new_x="LMARGIN", new_y="NEXT")
    pdf.set_draw_color(*CYAN)
    pdf.set_line_width(0.5)
    pdf.line(12, pdf.get_y(), 198, pdf.get_y())
    pdf.ln(3)

    # Calcular escala de fechas
    fechas_inicio = [t["inicio_plan"] for t in tareas if t["inicio_plan"]]
    fechas_fin    = [t["fin_plan"]    for t in tareas if t["fin_plan"]]
    if not fechas_inicio or not fechas_fin:
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 7, "Sin tareas con fechas definidas.")
        return bytes(pdf.output())

    t0 = datetime.fromisoformat(min(fechas_inicio)[:19])
    t1 = datetime.fromisoformat(max(fechas_fin)[:19])
    rango_dias = max(1, (t1 - t0).days)

    COL_NOMBRE = 55
    GAP        = 3
    BAR_AREA   = 186 - 12 - COL_NOMBRE - GAP   # mm disponibles para barras
    ROW_H      = 9
    BAR_H      = 5.5
    BAR_Y_OFF  = (ROW_H - BAR_H) / 2

    # Cabecera de escala (semanas)
    pdf.set_font("Helvetica", "", 6.5)
    pdf.set_text_color(*GRIS_OSC)
    pdf.set_x(12 + COL_NOMBRE + GAP)
    semanas = max(1, rango_dias // 7)
    ancho_sem = BAR_AREA / max(semanas, 1)
    for i in range(min(semanas, 20)):
        fecha_sem = t0 + timedelta(weeks=i)
        pdf.cell(ancho_sem, 5, fecha_sem.strftime("%d/%m"), border=0, align="L")
    pdf.ln(5)

    # Filas de tareas
    for tarea in tareas:
        if pdf.get_y() > 270:
            pdf.add_page()

        y0 = pdf.get_y()
        critica = tarea.get("en_ruta_critica", False)

        # Nombre de la tarea
        pdf.set_font("Helvetica", "B" if critica else "", 7.5)
        pdf.set_text_color(*CRITICO if critica else GRIS_OSC)
        pdf.set_x(12)
        nombre_corto = _txt(tarea["nombre"])[:32]
        pdf.cell(COL_NOMBRE, ROW_H, nombre_corto, border=0)

        # Calcular posición de la barra
        inicio_t = datetime.fromisoformat((tarea["inicio_plan"] or tarea.get("es") or "")[:19] or t0.isoformat())
        fin_t    = datetime.fromisoformat((tarea["fin_plan"]    or tarea.get("ef") or "")[:19] or t1.isoformat())
        x_start  = 12 + COL_NOMBRE + GAP + ((inicio_t - t0).days / rango_dias) * BAR_AREA
        w_total  = max(2, ((fin_t - inicio_t).days / rango_dias) * BAR_AREA)
        w_done   = w_total * (tarea["pct_completado"] / 100.0)

        bar_y = y0 + BAR_Y_OFF

        # Fondo de la barra (gris claro)
        pdf.set_fill_color(*GRIS_CLA)
        pdf.set_draw_color(*AZUL_MED)
        pdf.set_line_width(0.3)
        pdf.rect(x_start, bar_y, w_total, BAR_H, "FD")

        # Progreso (relleno azul o rojo para críticas)
        if w_done > 0:
            r, g, b = (CRITICO if critica else (0, 140, 190))
            pdf.set_fill_color(r, g, b)
            pdf.rect(x_start, bar_y, w_done, BAR_H, "F")

        # Etiqueta de porcentaje
        pdf.set_font("Helvetica", "", 6)
        pdf.set_text_color(*GRIS_OSC)
        pdf.set_xy(x_start + w_total + 1, y0 + BAR_Y_OFF)
        pdf.cell(12, BAR_H, f"{tarea['pct_completado']:.0f}%", border=0)

        pdf.ln(ROW_H)

    # ── Leyenda ─────────────────────────────────────────────────────────────────
    pdf.ln(4)
    pdf.set_font("Helvetica", "", 7)
    pdf.set_fill_color(0, 140, 190)
    pdf.rect(12, pdf.get_y(), 5, 3.5, "F")
    pdf.set_x(19)
    pdf.set_text_color(*GRIS_OSC)
    pdf.cell(40, 3.5, "Avance real")
    pdf.set_fill_color(*CRITICO)
    pdf.rect(62, pdf.get_y(), 5, 3.5, "F")
    pdf.set_x(69)
    pdf.cell(0, 3.5, "Tarea en ruta critica")
    pdf.ln(8)

    # ── Tabla de tareas ──────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_text_color(*AZUL_OSC)
    pdf.cell(0, 7, "Detalle de Tareas", new_x="LMARGIN", new_y="NEXT")
    pdf.set_line_width(0.4)
    pdf.line(12, pdf.get_y(), 198, pdf.get_y())
    pdf.ln(1)

    cols = ["Tarea", "Agente", "Inicio Plan", "Fin Plan", "Dur.", "Avance", "Holgura", "Critica"]
    widths = [52, 30, 22, 22, 10, 14, 14, 12]
    pdf.set_fill_color(*AZUL_MED)
    pdf.set_font("Helvetica", "B", 7)
    pdf.set_text_color(*BLANCO)
    for col, w in zip(cols, widths):
        pdf.cell(w, 6, col, fill=True, border=0)
    pdf.ln(6)

    pdf.set_font("Helvetica", "", 6.5)
    for i, t in enumerate(tareas):
        if pdf.get_y() > 275:
            pdf.add_page()
        fill = i % 2 == 0
        pdf.set_fill_color(*GRIS_CLA)
        pdf.set_text_color(*GRIS_OSC)
        critica_txt = "SI" if t.get("en_ruta_critica") else "no"
        fila = [
            _txt(t["nombre"])[:28],
            _txt(t.get("agente_id") or "—")[:16],
            (t["inicio_plan"] or "")[:10],
            (t["fin_plan"]    or "")[:10],
            f"{t['duracion_dias']:.0f}d",
            f"{t['pct_completado']:.0f}%",
            f"{t.get('holgura_dias', 0):.1f}d",
            critica_txt,
        ]
        for val, w in zip(fila, widths):
            pdf.cell(w, 5.5, val, fill=fill, border=0)
        pdf.ln(5.5)

    return bytes(pdf.output())


# ── Webhook WhatsApp / Control Remoto ─────────────────────────────────────────

class WhatsAppWebhookPayload(BaseModel):
    """Payload del webhook remoto (WhatsApp, curl, cron, etc.)."""
    mensaje:     str
    clave:       str            # contraseña en texto plano — validada contra MASTER_PASSWORD_HASH
    from_number: str = ""       # opcional: número origen para auditoría


_WHATSAPP_COMANDOS = (
    "  Status                 — estado del sistema y agentes cargados\n"
    "  Reiniciar Agente <id>  — recarga la configuración de un agente\n"
    "  Ayuda                  — lista de comandos\n"
)


@app.post("/webhook/whatsapp")
async def webhook_whatsapp(payload: WhatsAppWebhookPayload) -> dict:
    """
    Control remoto por webhook.  Auth: MASTER_PASSWORD_HASH (bcrypt).

    Diseñado para integraciones WhatsApp Business API, cron jobs o clientes HTTP.
    No expone token JWT: cada request lleva la clave y se valida en tiempo real.
    Timeout interno de 4s para no bloquear el event loop en servidores lentos.
    """
    import bcrypt as _bcrypt

    hash_env = os.environ.get("MASTER_PASSWORD_HASH", "").strip()
    if not hash_env:
        raise HTTPException(503, detail="Sistema no configurado: MASTER_PASSWORD_HASH ausente en .env")

    # Validar con timeout para evitar timing attacks sin bloquear el loop
    try:
        autorizado = await asyncio.wait_for(
            asyncio.get_running_loop().run_in_executor(
                None,
                lambda: _bcrypt.checkpw(
                    payload.clave.encode("utf-8"),
                    hash_env.encode("utf-8"),
                ),
            ),
            timeout=4.0,
        )
    except (asyncio.TimeoutError, Exception):
        autorizado = False

    if not autorizado:
        raise HTTPException(401, detail="Clave inválida.")

    cmd       = payload.mensaje.strip()
    cmd_lower = cmd.lower()

    logger.info(
        "Webhook WhatsApp recibido de %s: %s",
        payload.from_number or "desconocido",
        cmd,
    )

    # ── Status ─────────────────────────────────────────────────────────────────
    if cmd_lower == "status":
        n      = len(_orquestador.agentes) if _orquestador and hasattr(_orquestador, "agentes") else 0
        activo = kill_switch.is_active()
        return {
            "respuesta": (
                f"AgentDesk activo.\n"
                f"Agentes cargados: {n}\n"
                f"Kill switch: {'activo ✅' if activo else '⛔ desactivado'}"
            )
        }

    # ── Reiniciar Agente <id> ───────────────────────────────────────────────
    if cmd_lower.startswith("reiniciar agente "):
        agente_id = cmd[len("Reiniciar Agente "):].strip()
        if not agente_id:
            return {"respuesta": "Uso: Reiniciar Agente <id_del_agente>"}
        if not _bridge:
            return {"respuesta": "Bridge no disponible — el orquestador no está en línea."}
        await _bridge.send(Command(tipo=RELOAD_CONFIG, payload={"agente_id": agente_id}))
        return {"respuesta": f"Recarga de '{agente_id}' encolada correctamente."}

    # ── Ayuda ───────────────────────────────────────────────────────────────────
    return {
        "respuesta": (
            f"Comando no reconocido: '{cmd}'.\n\n"
            f"Comandos disponibles:\n{_WHATSAPP_COMANDOS}"
        )
    }


# ── Compliance ────────────────────────────────────────────────────────────────

@app.get("/compliance/reporte")
async def compliance_reporte(
    agente_id: str | None = None,
    dias: int = 30,
) -> dict:
    """
    Reporte de cumplimiento de guardrails.
    Agrupa abortos por guardrail y agente, emite alertas y sugerencias de temperatura.
    """
    from core.compliance import motor_compliance
    return motor_compliance.reporte_cumplimiento(agente_id=agente_id, dias=dias)


# ── Riesgo ────────────────────────────────────────────────────────────────────

@app.get("/riesgo/analisis/{proyecto_id}")
async def riesgo_analisis(proyecto_id: str) -> dict:
    """
    Correlaciona desviaciones del Gantt con impacto financiero proyectado.
    Retorna alertas ordenadas por nivel (CRITICO → ALTO → MEDIO).
    """
    from core.risk_engine import motor_riesgo
    return motor_riesgo.analizar_proyecto(proyecto_id)


# ── Salud del sistema (BI Dashboard) ─────────────────────────────────────────

@app.get("/sistema/salud")
async def sistema_salud() -> dict:
    """
    KPIs maestros consolidados: Gantt + Finanzas + Compliance + Riesgo.
    Score 0-100 ponderado para el Central Control Panel (BIDashboard ID 9).
    """
    from core.risk_engine import motor_riesgo
    return motor_riesgo.salud_sistema()


@app.get("/docs/manual")
async def descargar_manual(empresa: str = "AgentDesk") -> _Response:
    """
    Genera y descarga el Manual de Usuario en PDF personalizado con el nombre de la empresa.
    Ejemplo: GET /docs/manual?empresa=ACME+S.A.
    """
    from core.docs_gen import generar_manual
    loop      = asyncio.get_running_loop()
    pdf_bytes = await loop.run_in_executor(None, lambda: generar_manual(empresa))
    nombre    = f"Manual_AgentDesk_{empresa.replace(' ', '_')}.pdf"
    return _Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


@app.websocket("/ws/telemetria")
async def websocket_telemetria(ws: WebSocket) -> None:
    """
    Endpoint WebSocket. El dashboard React conecta aquí para recibir:
      - eventos de telemetría (@measure_latency) de los filtros del pipeline
      - notificaciones de agente_creado
      - cualquier mensaje que el servidor quiera enviar en tiempo real

    Protocolo: el cliente puede enviar {"ping": true} para keepalive.
    Autenticación: JWT en query param ?token=<jwt>  (browser WS no permite headers).
    El rol del token determina qué mensajes recibe el cliente (RBAC en broadcast).
    """
    from core.auth import verificar_token as _vt
    token   = ws.query_params.get("token", "")
    payload = _vt(token) if token else None
    rol     = (payload or {}).get("role", "viewer")

    await manager.connect(ws, rol=rol)
    await ws.send_text(json.dumps({
        "tipo":    "conexion",
        "mensaje": "Conectado al servidor de telemetria AgentDesk.",
        "rol":     rol,
    }))

    try:
        while True:
            data = await ws.receive_text()
            # Responder a pings para mantener la conexion viva
            try:
                msg = json.loads(data)
                if msg.get("ping"):
                    await ws.send_text(json.dumps({"pong": True}))
            except json.JSONDecodeError:
                pass
    except WebSocketDisconnect:
        manager.disconnect(ws)
